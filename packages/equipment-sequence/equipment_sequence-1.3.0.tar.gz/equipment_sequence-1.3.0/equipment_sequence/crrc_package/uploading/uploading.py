import datetime
import json
import time
from typing import List, Dict, Any

from bar_tender_api.bar_tender_api import BarTenderApi
from gkg_laser.laser_command import LaserCommand
from gkg_laser.laser_controller import LaserController
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from passive_equipment.handler_passive import HandlerPassive
from webservice_api.webservice_api import WebserviceAPI

from crrc_package.uploading.database_table_model import ProductIn, OriginMapData, ProductInStationLeft, \
    ProductInStationRight, MapDataInfo, CurrentLotInfo, NgProduct, CurrentOriginPath, ProductLabelInfo


class Uploading(HandlerPassive):

    def __init__(self):
        self.zebra = BarTenderApi(r"C:\Program Files\Seagull\BarTender Suite\SDK\Assemblies\Seagull.BarTender.Print.dll")
        time.sleep(3)
        super().__init__()
        self.template_dir = "D:/标签打印模板/"
        self.webservice_api = WebserviceAPI("http://10.96.141.69:7456/M1AutoPackageNotLogo.asmx?wsdl")
        self.gkg_laser = LaserController("127.0.0.1", 7005)
        self.gkg_laser.logger.addHandler(self.file_handler)

    def _get_product_in_info(self) -> dict:
        """获取进站产品信息.

        Returns:
            dict: 返回进站产品信息.
        """
        info = {
            "product_code": self.get_dv_value_with_name("product_code_in"),
            "state": self.get_dv_value_with_name("product_in_state_database")
        }
        return info

    def _get_product_in_station_left_info(self) -> dict:
        """获取产品放入左侧工位信息.

        Returns:
            dict: 返回产品放入左侧工位信息.
        """
        info = {
            "product_code": self.get_dv_value_with_name("product_code_in_station_left"),
            "station_index": self.get_dv_value_with_name("product_in_station_index_left")
        }
        return info

    def _get_product_in_station_right_info(self) -> dict:
        """获取产品放入右侧工位信息.

        Returns:
            dict: 返回产品放入右侧工位信息.
        """
        info = {
            "product_code": self.get_dv_value_with_name("product_code_in_station_right"),
            "station_index": self.get_dv_value_with_name("product_in_station_index_right"),
        }
        return info

    def product_in_station_left(self, call_back: dict):
        """左侧工位放入了产品, 保存产品信息.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])
        self.mysql.add_data(ProductInStationLeft, self._get_product_in_station_left_info())

    def product_in_station_right(self, call_back: dict):
        """右侧工位放入了产品, 保存产品信息.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])
        self.mysql.add_data(ProductInStationRight, self._get_product_in_station_right_info())

    @staticmethod
    def _add_group_index(group: list, group_index: int) -> list:
        """向分组里添加 group_index 字段.

        Args:
            group: 分组列表.
            group_index: 第几组.

        Returns:
            list: 添加 group_index 后的分组.
        """
        for _ in group:
            _.update({"group_index": group_index})
        return group

    def _group_data(self, data: List[Dict[str, Any]], group_size: int) -> List[List[Dict[str, Any]]]:
        """根据规则对数据进行分组。

        1. 每组中任意两个数据项的 vge_upper、vge_lower、vce_upper、vce_lower、vf_upper、vf_lower 的差值绝对值必须小于 0.1.
        2. 每组内的数据项按 index 值排序, 确保 index 差值最小.
        3. 每组的目标大小为 group_size，不满足的单独成组。

        Args:
            data: 数据列表, 每个元素是一个字典, 包含 vge_upper, vge_lower, vce_upper, vce_lower, vf_upper, vf_lower, index 等字段。
            group_size: 每组的目标大小。

        Returns:
            List[List[Dict[str, Any]]]: 分组后的数据列表，每个子列表代表一个组。
        """
        groups = []  # 存储最终的分组结果
        # 遍历数据，尝试将每个数据项分配到合适的组
        for item in data:
            placed = False  # 标记是否已分配到组
            # 尝试将当前数据项分配到已有的组
            for group in groups:
                # 检查当前数据项是否满足与组内所有数据项的规则
                if all(
                        abs(item["vge_upper"] - existing["vge_upper"]) < 0.1 and
                        abs(item["vge_lower"] - existing["vge_lower"]) < 0.1 and
                        abs(item["vce_25_upper"] - existing["vce_25_upper"]) < 0.1 and
                        abs(item["vce_25_lower"] - existing["vce_25_lower"]) < 0.1 and
                        abs(item["vf_25_upper"] - existing["vf_25_upper"]) < 0.1 and
                        abs(item["vf_25_lower"] - existing["vf_25_lower"]) < 0.1
                        for existing in group
                ):
                    # 如果组未满，则添加到组
                    if len(group) < group_size:
                        group.append(item)
                        placed = True
                        break
            # 如果没有找到合适的组，则创建一个新组
            if not placed:
                groups.append([item])
        # 处理未分组的数据（不满足规则的数据）
        ungrouped = []
        for group in groups:
            if 1 < len(group) < group_size:  # 如果组中数据项少于目标大小且大于1，说明未满足规则
                ungrouped.extend(group)
                groups.remove(group)
        # 将未分组的数据单独作为一组
        if ungrouped:
            groups.append(ungrouped)

        ok_list = []
        not_map_list = []
        not_equal_size_list = []
        group_index = 0
        for group in groups:
            if len(group) == group_size:
                group_index += 1
                new_group = self._add_group_index(group, group_index)
                ok_list.extend(new_group)
            elif len(group) == 1:
                not_map_list.extend(group)
            else:
                not_equal_size_list.extend(group)

        not_map_group_index = 1001
        final_not_nap_list = not_equal_size_list + not_map_list
        for index, one_row, in enumerate(final_not_nap_list, 1):
            if index % group_size == 0:
                not_map_group_index += 1
            one_row.update({"group_index": not_map_group_index})

        real_list = ok_list + final_not_nap_list
        return real_list

    def product_in_request(self, call_back:dict):
        """获取产品是否可以进站状态.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])
        instance = self.mysql.query_data_one(CurrentLotInfo)
        lot_name = instance.as_dict()["lot_name"]
        self.set_sv_value_with_name("current_lot_name", lot_name)

        info = {
            "product_code": self.get_dv_value_with_name("product_code_in"),
            "state": 1
        }

        self.mysql.add_data(ProductIn, info)

    def product_in_request_bak(self, call_back:dict):
        """获取产品是否可以进站状态.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])
        product_code = self.get_dv_value_with_name("product_code_in")
        instance = self.mysql.query_data_one(CurrentLotInfo)
        lot_name = instance.as_dict()["lot_name"]
        self.set_sv_value_with_name("current_lot_name", lot_name)

        request_params = {
            "material": product_code,
            "resource": self.get_dv_value_with_name("equipment_name"),
            "orderNumber": lot_name
        }
        try:
            response = self.webservice_api.call_query_function("TrackIn", request_params)
            self.logger.info("eap 反馈是: %s", response)
            eap_reply_in_state = response.Result
        except Exception as e:
            self.logger.info("eap 请求进站失败: %s", str(e))
            eap_reply_in_state = False
        self.logger.info("eap 进站结果: %s", eap_reply_in_state)
        if eap_reply_in_state:
            instance = self.mysql.query_data_one(OriginMapData, product_code=product_code)
            if instance:
                ng_instance = self.mysql.query_data_one(NgProduct, product_code=product_code)
                if ng_instance:
                    self.set_dv_value_with_name("product_in_state", 2)
                    self.set_dv_value_with_name("product_in_state_database", 2)
                else:
                    self.set_dv_value_with_name("product_in_state", 1)
                    self.set_dv_value_with_name("product_in_state_database", 1)
            else:
                self.set_dv_value_with_name("product_in_state", 2)
                self.set_dv_value_with_name("product_in_state_database", 3)
        else:
            self.set_dv_value_with_name("product_in_state", 2)
            self.set_dv_value_with_name("product_in_state_database", 4)
            self.logger.info("eap 不允许进站")
        self.mysql.add_data(ProductIn, self._get_product_in_info())

    def _generate_new_excel(self, map_data_list: list):
        """生产新的配对表."""
        # 将匹配后的数据生成新的配对表
        upper_key_list = [
            "product_code", "ices_ua_upper", "ices_ma_upper", "vge_upper", "iges_plus_upper",
            "iges_minus_upper", "vce_25_upper", "vce_150_upper", "vf_25_upper", "vf_150_upper",
            "td_on_upper", "tr_upper", "ton_upper", "eon_upper", "td_off_upper", "tf_upper",
            "toff_upper", "eoff_upper", "qrr_upper", "lrr_upper", "erec_upper"
        ]
        lower_key_list = [
            "product_code", "ices_ua_lower", "ices_ma_lower", "vge_lower", "iges_plus_lower",
            "iges_minus_lower", "vce_25_lower", "vce_150_lower", "vf_25_lower", "vf_150_lower",
            "td_on_lower", "tr_lower", "ton_lower", "eon_lower", "td_off_lower", "tf_lower",
            "toff_lower", "eoff_lower", "qrr_lower", "lrr_lower", "erec_lower"
        ]

        multiple_row_data = []
        # 多行数据
        for i, map_data in enumerate(map_data_list, 1):
            first_row_data = []
            first_row_data.append(str(i).zfill(4))
            for key in upper_key_list:
                value = map_data.get(key)
                first_row_data.append(value)
            first_row_data.insert(2, "UP")
            first_row_data.append(map_data["group_index"])
            second_row_data = []
            second_row_data.append(str(i).zfill(4))
            for key in lower_key_list:
                value = map_data.get(key)
                second_row_data.append(value)
            second_row_data.insert(2, "LOW")
            second_row_data.append(map_data["group_index"])

            multiple_row_data.append(first_row_data)
            multiple_row_data.append(second_row_data)
        # 原始文件路径和目标文件路径
        instance = self.mysql.query_data_one(CurrentOriginPath)
        source_file = instance.as_dict()["path"]
        time_str = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        target_file = f"{source_file.split('.')[0]}_{time_str}.xlsx"
        # 加载原始工作簿
        workbook = load_workbook(source_file)
        sheet = workbook.active  # 获取活动工作表
        # 清除X列的合并单元格, 遍历所有合并单元格
        merged_cells = list(sheet.merged_cells.ranges)
        for merged_cell in merged_cells:
            # 如果合并单元格在X列（第24列）
            if merged_cell.min_col == 24 or merged_cell.max_col == 24:
                sheet.unmerge_cells(str(merged_cell))
        # 删除第 8 行之后的所有行
        sheet.delete_rows(9, sheet.max_row - 8)
        # 写入多行数据
        for row_index, row_data in enumerate(multiple_row_data, start=9):  # 从第 1 行开始
            for col_index, value in enumerate(row_data, start=1):  # 从第 1 列开始
                if col_index == 24:
                    sheet.cell(row=row_index, column=col_index, value=str(value).zfill(4))
                else:
                    sheet.cell(row=row_index, column=col_index, value=value)

        # 定义Arial字体样式
        arial_font = Font(name="Arial", size=11)  # 设置字体为Arial，字号为11
        # 定义居中对齐样式
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 合并X列, 获取X列的列号
        x_column = 24
        # 遍历X列，查找连续相同的值
        start_row = 9  # 从第二行开始，假设第一行是标题
        end_row = sheet.max_row
        current_value = None
        start_merge_row = None
        for row in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row, column=x_column).value
            if cell_value and cell_value == current_value:
                # 如果当前值与上一个值相同，继续记录
                continue
            else:
                # 如果当前值与上一个值不同，检查是否需要合并
                if start_merge_row is not None and row - 1 > start_merge_row:
                    # 合并单元格
                    merge_range = f"{get_column_letter(x_column)}{start_merge_row}:{get_column_letter(x_column)}{row - 1}"
                    sheet.merge_cells(merge_range)
                    # 设置合并后的单元格内容居中
                    sheet.cell(row=start_merge_row, column=x_column).alignment = center_alignment
                # 更新当前值和起始行
                current_value = cell_value
                start_merge_row = row
        # 处理最后一组相同的值
        if start_merge_row is not None and end_row > start_merge_row:
            merge_range = f"{get_column_letter(x_column)}{start_merge_row}:{get_column_letter(x_column)}{end_row}"
            sheet.merge_cells(merge_range)

        # 定义实线边框样式
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        # 遍历所有单元格并添加边框
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = arial_font  # 设置字体为Arial
        # 保存为新文件
        workbook.save(target_file)

    def group_products_left(self, call_back: dict):
        """从数据库获取当前左侧工位的数据进行分组.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])

        # 将当前工位里面的产品和原始表连接获取产品的所有信息
        station_info_list = self.mysql.query_join(ProductInStationLeft, OriginMapData, "product_code")
        map_num = station_info_list[0]["map_number"]
        self.set_dv_value_with_name("map_number", map_num)

        map_data_list = self._group_data(station_info_list, map_num)
        self._generate_new_excel(map_data_list)

        self.mysql.delete_all_data(ProductInStationLeft)
        key_list = [
            "product_code", "vge_upper", "vge_lower", "vce_25_upper", "vce_25_lower",
            "vf_25_upper", "vf_25_lower", "group_index", "station_index"
        ]
        data_list = []
        for data_info in map_data_list:
            data_dict = {}
            for key, value in data_info.items():
                if key in key_list:
                    data_dict.update({key: value})
            data_list.append(data_dict)

        self.mysql.delete_all_data(ProductInStationLeft)
        self.mysql.delete_all_data(MapDataInfo)
        self.mysql.add_data_multiple(MapDataInfo, data_list)
        map_index_list = [_["station_index"] for _ in map_data_list]
        self.set_dv_value_with_name("map_info", map_index_list)

    def group_products_right(self, call_back: dict):
        """从数据库获取当前右边侧工位的数据进行分组.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])

        # 将当前工位里面的产品和原始表连接获取产品的所有信息
        station_info_list = self.mysql.query_join(ProductInStationRight, OriginMapData, "product_code")
        map_num = station_info_list[0]["map_number"]
        self.set_dv_value_with_name("map_number", map_num)
        map_data_list = self._group_data(station_info_list, map_num)
        self._generate_new_excel(map_data_list)

        self.mysql.delete_all_data(ProductInStationRight)
        key_list = [
            "product_code", "vge_upper", "vge_lower", "vce_25_upper", "vce_25_lower",
            "vf_25_upper", "vf_25_lower", "group_index", "station_index"
        ]
        data_list = []
        for data_info in map_data_list:
            data_dict = {}
            for key, value in data_info.items():
                if key in key_list:
                    data_dict.update({key: value})
            data_list.append(data_dict)
        self.mysql.delete_all_data(MapDataInfo)
        self.mysql.add_data_multiple(MapDataInfo, data_list)
        map_index_list = [_["station_index"] for _ in map_data_list]
        self.set_dv_value_with_name("map_info", map_index_list)

    def _get_label_info(self):
        """获取要进行打模块码的数据."""
        product_code = self.get_dv_value_with_name("product_code_label_print_request")
        current_lot_name = product_code[0:8]
        request_params = {"input": product_code, "orderNumber": current_lot_name}
        response = self.webservice_api.call_query_function("M1AutoPackagePrintInfo", request_params)
        self.logger.info("eap 返回的模块标签码信息: %s", response)
        label_code_type = response.Style
        if label_code_type == "99100":
            # self._print_99100_label(response.Context)
            self.set_dv_value_with_name("label_print_state", 1)
        elif label_code_type in ["阳光", "外销"]:
            self._print_static_label()
            self.set_dv_value_with_name("label_print_state", 1)
        else:
            self.set_dv_value_with_name("label_print_state", 2)

    def _print_99100_label(self, label_info):
        """打印99100模块码."""
        label_info_dict = json.loads(label_info)
        pn = label_info_dict.get("PN")
        sn = label_info_dict.get("SN")
        ssn = label_info_dict.get("SSN")
        mn = label_info_dict.get("MN")
        pd_type = label_info_dict.get("PDType")
        bar_code = label_info_dict.get("BarCode")
        self.set_dv_value_with_name("label_print_info", bar_code)
        self.zebra.execute_print(f"{self.template_dir}/static.btw")
        self.set_dv_value_with_name("label_print_state", 1)

    def _print_static_label(self):
        """打印阳光模块码."""
        product_code = self.get_dv_value_with_name("product_code_label_print_request")

        # 先确认是否是重复打标
        instance = self.mysql.query_data_one(ProductLabelInfo, product_code=product_code)
        if instance:
            label_info = instance.as_dict()
            update_data = {
                "product_index": label_info.get("product_index", ""),
                "first_row_info": label_info.get("first_row_info", ""),
                "second_row_info": label_info.get("second_row_info", ""),
                "third_row_info": product_code,
                "bar_code": label_info.get("bar_code", "")
            }
            self.zebra.execute_print(
                f"{self.template_dir}/static.btw",
                update_data=update_data, close_btw=False, close_engine=False
            )
            self.set_dv_value_with_name("label_print_state", 1)
            self.set_dv_value_with_name("label_print_info", label_info.get("bar_code", ""))
            return

        if self.get_dv_value_with_name("current_sun_index") == 9999:
            self.set_dv_value_with_name("current_sun_index", 1)
            self.config_instance.update_config_dv_value("current_sun_index", 1)

        current_sun_index = self.get_dv_value_with_name("current_sun_index")
        current_sun_index_str = str(current_sun_index).zfill(4)
        instance = self.mysql.query_data_one(OriginMapData, product_code=product_code)
        product_info = instance.as_dict()
        uppers_str = f"{product_info['vge_upper']:.2f}  {product_info['vce_25_upper']:.2f}  {product_info['vf_25_upper']:.2f}"
        lowers_str = f"{product_info['vge_lower']:.2f}  {product_info['vce_25_lower']:.2f}  {product_info['vf_25_lower']:.2f}"
        vge = f"{product_info['vge_upper']:.2f} {product_info['vge_lower']:.2f}"
        vce = f"{product_info['vce_25_upper']:.2f} {product_info['vce_25_lower']:.2f}"
        vf = f"{product_info['vf_25_upper']:.2f} {product_info['vf_25_lower']:.2f}"
        bar_code = f"{product_code} {vge} {vce} {vf}"

        # 将要打印的信息保存到数据库
        insert_data = {
            "product_code": product_code,
            "product_index": current_sun_index_str,
            "first_row_info": uppers_str,
            "second_row_info": lowers_str,
            "bar_code": bar_code
        }
        self.mysql.add_data(ProductLabelInfo, insert_data)

        update_data = {
            "product_index": current_sun_index_str,
            "first_row_info": uppers_str,
            "second_row_info": lowers_str,
            "third_row_info": product_code,
            "bar_code": bar_code
        }
        self.zebra.execute_print(
            f"{self.template_dir}/static.btw",
            update_data=update_data, close_btw=False, close_engine=False
        )
        self.set_dv_value_with_name("label_print_state", 1)
        self.set_dv_value_with_name("label_print_info", bar_code)
        self.set_dv_value_with_name("current_sun_index", current_sun_index + 1)
        self.config_instance.update_config_dv_value("current_sun_index", current_sun_index + 1)

    def label_print_request(self, call_back: dict):
        """产品请求打标签码.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])
        # 调用请求打那种码的接口
        self._get_label_info()

        self.set_dv_value_with_name("label_print_state", 1)

    def laser_print_request(self, call_back: dict):
        """产品请求激光打码.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])

        instance = self.mysql.query_data_one(OriginMapData)
        is_laser_print = instance.as_dict()["is_laser_print"]
        laser_type = instance.as_dict()["laser_type"]

        self.set_dv_value_with_name("is_laser_print", is_laser_print)
        self.set_dv_value_with_name("laser_logo_type", laser_type)
        product_type = self.mysql.query_data_one(OriginMapData).as_dict()["product_type"]
        self.set_dv_value_with_name("product_type", product_type)

    def laser_print(self, call_back: dict):
        """开始打印激光码.

        Args:
            call_back: 要执行的 call_back 信息.
        """
        self.logger.info("正在执行 %s 函数", call_back["operation_func"])
        laser_type_map = {
            1: "M1_C_TG", 2: "M1_D_TG"
        }
        laser_type = self.mysql.query_data_one(OriginMapData).as_dict()["laser_type"]
        self.logger.info("数据库记录要打激光的类型是: %s", laser_type)
        file_name = laser_type_map[laser_type]
        self.logger.info("对应的模板文件是: %s", file_name)

        state = self.gkg_laser.execute_command(LaserCommand.INITIALIZE, file_name)
        self.logger.info("激光软件是否打开: %s", state)
        if state:
            if laser_type == 1:
                product_type_data_base = self.mysql.query_data_one(OriginMapData).as_dict()["product_type"]
                self.logger.info("数据库产品码: %s", product_type_data_base)
                product_type = self.get_dv_value_with_name("product_type")
                self.logger.info("当前产品码: %s", product_type)
                if product_type_data_base != product_type:
                    self.logger.info("更新产品码: %s", product_type)
                    self.gkg_laser.execute_command(LaserCommand.DATA_UPDATE, product_type)
                else:
                    self.logger.info("不需要更新产品码: %s", product_type)

            self.gkg_laser.execute_command(LaserCommand.MARK_START)
            self.wait_time(4)
            self.set_dv_value_with_name("laser_success", 1)
        else:
            self.set_dv_value_with_name("laser_success", 2)
