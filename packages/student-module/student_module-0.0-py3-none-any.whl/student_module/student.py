import mysql.connector
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import datetime
import re

class Student():
    def __init__(self):
        self.get_details()
 
    def get_details(self):
        while True:
            self.name=input("Enter name :")
            if (self.name.isalpha()):
                break
            else:
                print("Invalid name")

        while True:
            self.rollNo=input("Enter roll number :")
            if (self.rollNo.isdigit() and len(self.rollNo) == 5):
                break
            else:
                print("Invalid roll number")

        while True: 
            self.DOB=input("Enter DOB 'YYYY-MM-DD' :")
            try:
                datetime.datetime.strptime(self.DOB, "%Y-%m-%d")
                break
            except ValueError:
                print("Invalid DOB format. Use YYYY-MM-DD.")

            
        while True:
            self.mobileNo=input("Enter mobile number :")
            if (self.mobileNo.isdigit() and len(self.mobileNo) == 10):
                break
            else:
                print("Invalid roll number")

        while True:
            self.email=input("Enter email :")
            email_pattern = r'^[a-zA-Z0-9._]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if re.match(email_pattern, self.email):
                break
            else:
                print("Invalid email address.")
           
        self.marks=[]
        for subject in ["Tamil","English","Maths","Science","Social"]:
            while True:
                mark=int(input("Enter "+ subject +" mark :"))
                if (mark >=0 and mark <=100):
                    self.marks.append(mark)
                    break
                else:
                    print("Invalid mark")

    def display(self):
        print("Name :"+self.name)
        print("Roll number: "+self.rollNo)
        print("Date of birth :" +self.DOB)
        print("Mobile number; " +self.mobileNo)
        print("Email: " +self.email)

        self.total=0
        for i in self.marks:
            self.total=self.total+i

        print("Total :", self.total)

        self.average=self.total/5
        print("Average: " , self.average)

        self.percentage=(self.total/500)*100
        print("Percentage: " ,self.percentage)
        

    def txtFile(self):
        folder=("D:\\Python\\pythonPractice")
        file=self.name+self.rollNo
        fileName= os.path.join(folder, file + ".txt")
        

        T_mark, E_mark, M_mark, S_mark, SS_mark = self.marks

        with open (fileName ,"a") as f:
            f.write("Student name : " + self.name + "\n")
            f.write("Roll number : " + self.rollNo + "\n")
            f.write("DOF : " + self.DOB + "\n")
            f.write("Mobil number : " + self.mobileNo + "\n")
            f.write("Email : " + self.email + "\n")
            f.write("Tamil mark : " + str(T_mark) + "\n")
            f.write("English mark : " + str(E_mark) + "\n")
            f.write("Maths mark : " + str(M_mark) + "\n")
            f.write("Science mark : " + str(S_mark) + "\n")
            f.write("Social sciencemark : " + str(SS_mark)+ "\n")
            f.write("Total: " + str(self.total)+ "\n" )
            f.write("Average: " + str(self.average)+ "\n")
            f.write("Percentage: " + str(self.percentage)+ "\n")
        print("Data successfully added to txt file")

    def db(self):
        mydb=mysql.connector.connect(
            host="localhost",
            user="root",
            password="patterns",
            database="student"
            )
        
        mycursor=mydb.cursor()

        T_mark, E_mark, M_mark, S_mark, SS_mark = self.marks

        mycursor.execute("CREATE TABLE IF NOT EXISTS details(ID INT AUTO_INCREMENT PRIMARY KEY,RollNo INT,Name VARCHAR(250),DOB DATE,MobileNo VARCHAR(10),Email VARCHAR(255),T_mark INT,E_mark INT,M_mark INT,S_mark INT,SS_mark INT,Total INT,Average FLOAT,Percentage FLOAT)") 
 
        sql = "insert into details(RollNo,Name,DOB,MobileNo,Email,T_mark,E_mark,M_mark,S_mark,SS_mark,Total,Average,Percentage)VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    
        val = (int(self.rollNo), self.name, self.DOB, self.mobileNo, self.email,T_mark, E_mark,
               M_mark, S_mark, SS_mark,self.total, self.average, self.percentage)


        mycursor.execute(sql,val)
        mydb.commit()
        print("Table created and data added successfully")

    

    def excel(self):
        T_mark, E_mark, M_mark, S_mark, SS_mark = self.marks

        folder_path = "D:\\Python\\pythonPractice"
        file_path = os.path.join(folder_path, "TestExcel.xlsx")

        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append([
                "Roll No", "Name", "DOB", "Mobile No", "Email",
                "Tamil", "English", "Maths", "Science", "Social",
                "Total", "Average", "Percentage"
            ])
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="80FF00", end_color="80FF00",fill_type = "solid")

        sheet.append([
           self.rollNo, self.name, str(self.DOB), self.mobileNo, self.email,
           T_mark, E_mark, M_mark, S_mark, SS_mark,self.total,
           round(self.average, 2), round(self.percentage, 2)
        ])

        wb.save(file_path)
        print("Student details saved to Excel")

        
        

        

        
