# src/file_downloader/server.py
import os, urllib.parse
import httpx
from fastmcp import FastMCP
import openpyxl
import shutil
import json

def main():
    mcp = FastMCP("excel-mcp-stdio")

    @mcp.tool()
    def write_data_to_excel_with_custom(
        filepath: str,
        outputpath: str,
        data: dict,
    ) -> str:
        """
        將資料寫入 Excel 模板中對應的 {{key}} 位置。

        範例:
        write_data_to_excel_with_custom(
            filepath="D:/excel_temp.xlsx",
            outputpath="D:/excel_filled.xlsx",
            data={"item": "產品A", "number": "2", "price": 10000}
        )
        """
        try:
            params = {
                "filepath": filepath,
                "outputpath": outputpath,
                "data": data
            }
            return fill_excel(params)
        except Exception as e:
            return f"Error: {e}"


    @mcp.tool()
    def write_data_to_excel_with_insert_row(
        data: list[list],
    ) -> str:
        """
        將資料寫入 Excel 模板中
        """
        try:
            params = {
                "filepath": "D:/excel_auto_temp.xlsx",
                "outputpath": "D:/excel__auto_2025.xlsx",
                "sheet_name": "英文翻譯",
                "data_rows": data
            }
            return fill_excel_with_insert_row(params)
        except Exception as e:
            return f"Error: {e}"

    mcp.run(transport="stdio")


def fill_excel(params: dict) -> str:
    src = params["filepath"]
    dst = params["outputpath"]
    data = params["data"]

    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    sheet_name = "英文翻譯"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到名稱為 \"{sheet_name}\" 的工作表")
    ws = wb[sheet_name]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                for key, val in data.items():
                    cell.value = cell.value.replace(f"{{{{{key}}}}}", str(val))

    wb.save(dst)
    return dst


def fill_excel_with_insert_row(params: dict) -> str:
    src = params["filepath"]
    dst = params["outputpath"]
    sheet_name = params["sheet_name"]
    data = params["data_rows"]

    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到名稱為 \"{sheet_name}\" 的工作表")
    ws = wb[sheet_name]

    start_col = 3   # C 欄
    start_row = 20

    for i, row_vals in enumerate(data):
        target_row = start_row + i
        if i > 0:                      # 第二筆起，先插入一列
            ws.insert_rows(target_row)

        for j, val in enumerate(row_vals):
            ws.cell(row=target_row, column=start_col + j, value=val)

    wb.save(dst)
    return dst







if __name__ == "__main__":
    main()
