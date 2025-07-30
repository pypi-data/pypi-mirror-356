import openpyxl
import shutil
import json
import os

def fill_excel(params: dict) -> str:
    """
    params = {
      "filepath": ".../excel_temp.xlsx",
      "outputpath": ".../excel_2025.xlsx",
      "data": {"item": "...", "number": "...", "price": ...}
    }
    """
    src = params["filepath"]
    dst = params["outputpath"]
    data = params["data"]

    # 确保含目录
    os.makedirs(os.path.dirname(dst), exist_ok=True)

    # 复制模板
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    sheet_name = "英文翻譯"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到名稱為 \"{sheet_name}\" 的工作表")
    ws = wb[sheet_name]

    # 遍歷替換 {{key}} 內容
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and "{{" in val and "}}" in val:
                text = val
                for key, v in data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in text:
                        text = text.replace(placeholder, str(v))
                cell.value = text

    wb.save(dst)
    return dst

# # 以下範例展示如何調用 fill_excel(params)：
# if __name__ == "__main__":
#     params = {
#         "filepath": "D:/excel_temp.xlsx",
#         "outputpath": "D:/excel_2025.xlsx",
#         "data": {"item": "產品A", "number": "2", "price": 10000}
#     }
#     out = fill_excel(params)
#     print(f"已生成：{out}")