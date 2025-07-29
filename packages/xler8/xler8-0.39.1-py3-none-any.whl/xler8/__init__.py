import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import subprocess
import os
import pandas
import copy


def finder_expose(filename_or_path):
    subprocess.call("""/bin/bash -c 'open -R "%s"'""" % filename_or_path, shell=True)




def xlsx_sheetnames(filename:str):
    with pandas.ExcelFile(filename) as f:
        return(list(f.sheet_names))

def xlsx_rows_row_callback__r__hdr__row__row_assoc__row_named__context():
    pass

def xlsx_rows(filename:str, sheetname:str, row_callback, context):
    """row_callback(r, hdr, row, row_assoc, row_named, context).
    """
    df = pandas.read_excel(filename, dtype=str, header=None, sheet_name=sheetname, keep_default_na=False)
    df.fillna('')
    n_rows = len(df)
    n_cols = len(df.columns)

    hdr = []
    for col in range(0, n_cols):
        hdr.append(df.loc[0,col])

    # skip row0
    for r in range(1, n_rows):
        row = []
        row_assoc = {}
        row_named = {}
        for col in range(0, n_cols):
            row.append(df.loc[r,col])
            row_assoc[hdr[col]] = df.loc[r,col]
            row_named[get_column_letter(1+col)] = df.loc[r,col]
        row_callback(r+1, hdr, row, row_assoc, row_named, context)


def out(filename:str, sheets:dict):
    xlsx_out(filename=filename, sheets=sheets)


def xlsx_out(filename:str, sheets:dict):
    """Simple Excel Writer.


    mxee.helper.xlsx_out(
        "test2.xlsx",
        sheets={
            'blatt1': {
                'data': [['A', 'B', 'C'],['120']],
                'cw':{'A': 50}
            }
        }
    )
"""
    try:
        with pd.ExcelWriter(filename) as ew:

            for k in sheets.keys():
                dat = sheets[k]['data']

                dfc = pd.DataFrame(dat) # dataframe current
                dfc.to_excel(ew, sheet_name=k, index=False, header=False)
            
                col_widths = {}
                if 'cw' in sheets[k].keys():
                    col_widths = sheets[k]['cw']

                for wdef_k in col_widths.keys():
                    ew.sheets[k].column_dimensions[wdef_k].width = col_widths[wdef_k]

                for header0_idx in range(0, len(dat[0])):
                    column_letter = get_column_letter(header0_idx+1)
                    ew.sheets[k][column_letter + "1"].font = Font(bold=True)
                
                ew.sheets[k].freeze_panes = ew.sheets[k]['A2']

    except Exception as e:
        import os
        os.unlink(filename)
        print(e)


def xlsx_named_columns_callback(r, hdr, row, row_assoc, row_named, ctx):
    for h in hdr:
        if not h in ctx.keys():
            ctx[h] = []

    for h in hdr:
        ctx[h].append(row_assoc[h])

    # sheetname = context['sheet']
    # print(sheetname)
    # for h in hdr:
    #     if context['res']['sheetname']
    #     print(h)
    #     context['h'] = h


def xlsx_named_columns(filename):
    sheetnames = xlsx_sheetnames(filename=filename)
    data = {}
    for sheetname in sheetnames:
        ctx = {}
        xlsx_rows(filename, sheetname, xlsx_named_columns_callback, ctx)
        data[sheetname] = copy.deepcopy(ctx)
    return data


def xlsx_all_columns_by_sheet_and_header(filename):
    """Returns a dict of dicts of lists 'sheetname'->'column-header'->[all str values of column].
    """
    return xlsx_named_columns(filename)


def xlsx(filename):
    """Returns a dict of dicts of lists 'sheetname'->'column-header'->[all str values of column].
    """
    return xlsx_named_columns(filename)


def xlsx0(filename):
    """Returns a dict of lists 'column-header'->[all str values of column] of the first sheet.
    """
    sheetnames = xlsx_sheetnames(filename=filename)
    return xlsx_named_columns(filename)[sheetnames[0]]



def xlsx_all_headers_by_sheet__callback(r,hdr,row,row_assoc,row_named,context):
    if len(context['hdr']) == 0:
        context['hdr'] = copy.deepcopy(hdr)


def headers0(filename):
    sheetnames = xlsx_sheetnames(filename)
    return copy.deepcopy(xlsx_all_headers_by_sheet(filename)[sheetnames[0]])


def xlsx_all_headers_by_sheet(filename):
    sheetnames = xlsx_sheetnames(filename=filename)
    res={}
    for sheetname in sheetnames:
        ctx={'hdr': []}
        xlsx_rows(filename=filename, sheetname=sheetname, row_callback=xlsx_all_headers_by_sheet__callback, context=ctx)
        res[sheetname] = copy.deepcopy(ctx['hdr'])
    return res


def xlsx_column_info_ascii(filename):
    sheetnames = xlsx_sheetnames(filename=filename)
    hdr = xlsx_all_headers_by_sheet(filename=filename)
    lines = []
    for sheetname in sheetnames:
        for col_index in range(0, len(hdr[sheetname])):
            lines.append(sheetname + "\t" + get_column_letter(col_index+1) + "\t" + hdr[sheetname][col_index])
        lines.append("")
    return "\n".join(lines)


def transpose1(src_list):
    res=[]
    for x in src_list:
        res.append([x])
    return res

# def xlsx_all_ascii(filename):
#     """TODO.
#     """
#     sheetnames = xlsx_sheetnames(filename=filename)
#     data = xlsx_all_columns_by_sheet_and_header(filename)
#     res = []
#     for s in sheetnames:
#         sheet_data = data[s]
#     return ""


def sheet_lookup(sheet_data, search_column_name, search_value, target_column_name):
    res = []
    rowcount = len(sheet_data[search_column_name])
    for i in range(0, rowcount):
        if sheet_data[search_column_name][i] == search_value:
            res.append(sheet_data[target_column_name][i])
    return res


def sheet_match(sheet_data, match_column_name, match_values_or, only_headers=None):
    res = []
    rowcount = len(sheet_data[match_column_name])
    for i in range(0, rowcount):
        if sheet_data[match_column_name][i] == match_values_or[0]:
            nrow = []
            for oh in only_headers:
                nrow.append(sheet_data[oh][i])
            res.append(nrow)
    return res


def sheet_flatten(sheet_data, index_column_name, k_name, v_name, fieldnames, row_grep_i=""):
    groups_ = {}
    for x in sheet_data[index_column_name]:
        groups_[x]=True

    groups = list(sorted(groups_.keys()))

    res = [[index_column_name] + copy.deepcopy(fieldnames)]

    stacks = {}
    for f in fieldnames:
        stacks[f] = {}

    for i in range(0, len(sheet_data[index_column_name])):
        g = sheet_data[index_column_name][i]
        k = sheet_data[k_name][i]
        v = sheet_data[v_name][i]
        for f in fieldnames:
            if k == f:
                stacks[k][g] = v

    GREP_I = row_grep_i.upper()

    for g in groups:
        row = [g]
        for f in fieldnames:
            row.append(stacks[f].get(g, ""))
        if row_grep_i == "":
            res.append(row)
        else:
            ROWJOINED = " ".join(row)
            ROWJOINED = ROWJOINED.upper()
            if ROWJOINED.find(GREP_I) >= 0:
                res.append(row)

    return res



def cw_gen(data, colf=1.4):
    col_chars = []
    if len(data) > 0:
        col_chars = [ 0 for x in range(0, len(data[0])) ]
    
    for row in data:
        for rvi in range(0, len(row)):
            new_length = len(str(row[rvi]))
            if new_length > col_chars[rvi]:
                col_chars[rvi] = new_length

    cw={}
    for hci in range(0, len(col_chars)):
        cw[get_column_letter(1+hci)] = int(col_chars[hci] * colf)

    return cw
