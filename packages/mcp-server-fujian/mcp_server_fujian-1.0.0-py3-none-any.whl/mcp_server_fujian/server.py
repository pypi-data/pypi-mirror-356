import os
import zipfile
import shutil
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from enum import Enum
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent, INVALID_PARAMS, INTERNAL_ERROR
import logging

class FujianTools(str, Enum):
    PROCESS_FUJIAN = "process_fujian"


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('extract_zips.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def extract_zips(zip_file_path: str, output_dir: str) -> None:
    zip_count = 1
    tmp_dir = Path(output_dir) / 'tmp'
    output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    zip_path = Path(zip_file_path)
    if not zip_path.exists() or not zip_path.is_file() or not zip_path.suffix == '.zip':
        logger.error(f"{zip_path} 不是有效的 ZIP 文件")
        return
    
    logger.info(f"处理压缩包: {zip_path}")
    
    tmp_zip_dir = tmp_dir / f"zip_{zip_count}"
    tmp_zip_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zfile:
            for zip_file in zfile.namelist():
                try:
                    decoded_name = zip_file.encode('cp437').decode('gbk')
                except:
                    decoded_name = zip_file
                
                target_path = output_dir / decoded_name
                logger.info(f"处理压缩包中的文件：{decoded_name}")
                
                if target_path.exists():
                    logger.warning(f"文件已存在：{target_path}")
                    continue
                
                zfile.extract(zip_file, tmp_zip_dir)
                extracted_file_path = tmp_zip_dir / zip_file
                
                if decoded_name.endswith('.zip'):
                    logger.info(f"发现嵌套 ZIP 文件：{decoded_name}，开始递归解压")
                    try:
                        extract_zips(extracted_file_path, output_dir)
                        logger.info(f"嵌套 ZIP 文件解压完成：{decoded_name}")
                    except Exception as e:
                        logger.error(f"递归解压嵌套 ZIP 文件失败：{decoded_name}，错误：{e}")
                    continue
                
                if decoded_name.endswith(('.xlsx', '.xls')):
                    logger.info(f"处理 Excel 文件：{decoded_name}")
                    try:
                        wb = openpyxl.load_workbook(extracted_file_path)
                        wb.save(target_path)
                        logger.info(f"Excel 文件已保存，格式保留：{target_path}")
                    except Exception as e:
                        logger.error(f"处理 Excel 文件失败：{e}")
                        os.makedirs(target_path.parent, exist_ok=True)
                        os.rename(extracted_file_path, target_path)
                else:
                    os.makedirs(target_path.parent, exist_ok=True)
                    os.rename(extracted_file_path, target_path)
                
                logger.info(f"{zip_file} -> {target_path}")
    
    except Exception as e:
        logger.error(f"处理压缩包失败：{zip_path}，错误：{e}")
    
    if tmp_zip_dir.exists():
        shutil.rmtree(tmp_zip_dir)
        logger.info(f"已删除临时目录：{tmp_zip_dir}")
    
    logger.info(f"已处理第 {zip_count} 个压缩包")


def check_and_extract(folder_path, keywords=None, exclude_keyword='芜湖路'):
    if keywords is None:
        keywords = [
            '芜湖市供电公司', '镜湖区供电公司', '鸠江区供电公司',
            '弋江区供电公司', '湾沚区供电公司', '繁昌区供电公司',
            '南陵县供电公司', '无为市供电公司', '芜湖公司',
        ]
    
    extracted_files = {keyword: {'filtered': [], 'feedback': [], 'other': []} for keyword in keywords}
    has_non_feedback = False
    other_files = []
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            if not file.endswith(('.csv', '.xlsx', '.xls', '.zip')):
                print(f"记录其他文件：{file_path}")
                other_files.append(file_path)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(('.csv', '.xlsx', '.xls')):
                file_path = os.path.join(root, file)
                print(f"处理文件: {file}")

                if "反馈" not in file:
                    has_non_feedback = True
                    try:
                        df = pd.read_excel(file_path) if file.endswith(('.xlsx', '.xls')) else pd.read_csv(file_path)
                        df = df.astype(str)

                        mask_include = df.apply(lambda x: x.str.contains('|'.join(keywords), na=False)).any(axis=1)
                        mask_exclude = df.apply(lambda x: x.str.contains(exclude_keyword, na=False)).any(axis=1)
                        mask = mask_include & ~mask_exclude

                        if mask.any():
                            extracted_data = df[mask]
                            for keyword in keywords:
                                keyword_mask = extracted_data.apply(lambda x: x.str.contains(keyword, na=False)).any(axis=1)
                                if not keyword_mask.any():
                                    continue
                                new_file_path = os.path.join(root, f'extracted_{keyword}_{file}')
                                if file.endswith('.csv'):
                                    extracted_data[keyword_mask].to_csv(new_file_path, index=False, encoding='utf-8-sig')
                                else:
                                    wb = openpyxl.Workbook()
                                    ws = wb.active
                                    for col_num, column_title in enumerate(extracted_data[keyword_mask].columns, 1):
                                        ws.cell(row=1, column=col_num).value = column_title
                                    for row_num, row_data in enumerate(extracted_data[keyword_mask].values, 2):
                                        for col_num, cell_value in enumerate(row_data, 1):
                                            ws.cell(row=row_num, column=col_num).value = cell_value
                                    wb.save(new_file_path)
                                extracted_files[keyword]['filtered'].append(new_file_path)
                    except Exception as e:
                        print(f"处理文件 {file_path} 失败：{str(e)}")
                else:
                    print(f"记录反馈文件，保留格式：{file_path}")
                    for keyword in keywords:
                        extracted_files[keyword]['feedback'].append(file_path)

    for keyword in keywords:
        extracted_files[keyword]['other'] = other_files

    return extracted_files, has_non_feedback


def zip_files(folder_path, extracted_files, has_non_feedback):
    if has_non_feedback:
        for keyword, file_dict in extracted_files.items():
            if file_dict['filtered']:
                files = file_dict['filtered'] + file_dict['feedback'] + file_dict['other']
                if files:
                    zip_file_path = os.path.join(folder_path, f'{keyword}.zip')
                    try:
                        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            for file in files:
                                zipf.write(file, os.path.relpath(file, folder_path))
                                print(f"已添加文件到压缩包：{file}（关键词：{keyword}）")
                        print(f"已生成压缩文件：{zip_file_path}")
                    except Exception as e:
                        print(f"生成压缩文件 {zip_file_path} 失败：{str(e)}")
            else:
                print(f"未生成压缩文件：{keyword} 无筛选表格")
    else:
        print("所有表格均为反馈表，生成统一压缩包")
        feedback_files = list(set(extracted_files[list(extracted_files.keys())[0]]['feedback']))
        other_files = list(set(extracted_files[list(extracted_files.keys())[0]]['other']))
        files = feedback_files + other_files
        if files:
            zip_file_path = os.path.join(folder_path, 'all_feedback_and_others.zip')
            try:
                with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file in files:
                        zipf.write(file, os.path.relpath(file, folder_path))
                        print(f"已添加文件到压缩包：{file}（统一压缩包）")
                print(f"已生成统一压缩文件：{zip_file_path}")
            except Exception as e:
                print(f"生成统一压缩文件 {zip_file_path} 失败：{str(e)}")
        else:
            print("无反馈文件或其他文件，未生成统一压缩包")


def process_files(zip_file_path: str, output_dir: str) -> str:
    try:
       
        extract_zips(zip_file_path, output_dir)
        if not os.path.exists(output_dir):
            raise Exception("解压后的输出目录不存在")

        extracted_files, has_non_feedback = check_and_extract(output_dir)
        if extracted_files is None:
            raise Exception("表格文件处理失败，未提取任何文件")
        
        zip_files(output_dir, extracted_files, has_non_feedback)

        tmp_dir = Path(output_dir) / 'tmp'
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir)

        return "所有处理步骤已完成"
    except Exception as e:
        error_message = f"处理过程中发生错误：{str(e)}"
        return error_message


class FujianServer:
    def process_info_codes(self, info_code: str) -> str:
        """处理数据库查询结果中的 info_codes，调用 process_files 进行文件处理"""
        
        zip_file_path = f"D:/gzrw_file/gzrw_file/{info_code}.zip"
        output_dir = f"D:/gzrw_file/{info_code}"

        if os.path.exists(zip_file_path):
            result=process_files(zip_file_path, output_dir)
            return f"成功处理 info_code: {info_code}- {result}"
        else:
            return f"压缩包未找到：{zip_file_path}"

async def serve() -> None:
    server = Server("mcp-fujian")
    fujian_server = FujianServer()

    @server.list_tools()
    async def list_tools() -> list[Tool]:
        return [
            Tool(
                name=FujianTools.PROCESS_FUJIAN.value,
                description="根据数据编号对指定的附件内容进行解析，包括表格拆分和压缩。",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "info_code": {
                            "type": "str",
                            "description": "工单编号",
                        },
                    },
                    "required": ["info_code"]
                },
            ),
        ]

    @server.call_tool()
    async def call_tool(name: str, arguments: dict) -> list[TextContent]:
        try:
            match name:
                case FujianTools.PROCESS_FUJIAN.value:
                    info_code = arguments.get("info_code")
                    if not info_code:
                        raise ValueError("Missing required argument: info_code")
                    result = fujian_server.process_info_codes(info_code)

                case _:
                    raise ValueError(f"Unknown tool: {name}")
            return [TextContent(type="text", text=result)]

        except Exception as e:
            raise ValueError(f"Error processing mcp-server-fujian query: {str(e)}")

    options = server.create_initialization_options()
    async with stdio_server() as (read_stream, write_stream):
        await server.run(read_stream, write_stream, options, raise_exceptions=True)