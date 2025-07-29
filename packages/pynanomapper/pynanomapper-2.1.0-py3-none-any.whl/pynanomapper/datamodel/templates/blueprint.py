import pandas as pd
import os
import json
from datetime import datetime
from xlsxwriter.utility import xl_col_to_name
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl import load_workbook
from openpyxl.comments import Comment
from copy import copy


def iom_format(df, param_name="param_name", param_group="param_group"):
    df.fillna(" ", inplace=True)
    #print(df.columns)
    # Create a new DataFrame with one column
    result_df = pd.DataFrame(columns=['param_name'])
    result_df["type"] = "group"
    result_df["position"] = -1
    # Iterate through unique groups
    for group in df[param_group].unique():
        group_df = df[df[param_group] == group]
        # Get names for the group
        names = group_df[param_name].tolist()
        # Append group and names to the result DataFrame
        tmp = pd.DataFrame({'param_name': [group] + names })
        tmp["type"] = "names"
        tmp.at[0, "type"] = "group"
        tmp['position'] = -1
        result_df = pd.concat([result_df, tmp], ignore_index=True)
    return result_df


def json2frame(json_data, sortby=None):
    tmp = pd.DataFrame(json_data)
    if sortby is None:
        return tmp
    else:
        return tmp.sort_values(by=sortby)


def get_method_metadata(bp_json):
    _header = {
        "Project Work Package": bp_json.get("provenance_workpackage", ""),
        "Partner conducting test/assay": bp_json.get("provenance_provider", ""),
        "Test facility - Laboratory name": bp_json.get("provenance_provider", ""),
        "Lead Scientist & contact for test": bp_json.get("provenance_contact", ""),
        "Assay/Test work conducted by": bp_json.get("provenance_operator", ""),
        "Full name of test/assay": bp_json.get("METHOD", ""),
        "Short name or acronym for test/assay": bp_json.get("METHOD", ""),
        "Type or class of experimental test as used here": bp_json.get(
            "PROTOCOL_CATEGORY_CODE", ""),
        "End-Point being investigated/assessed by the test":  [
            item["result_name"] if "result_name" in item else "result_name_not_specified" for item in bp_json.get("question3", [])
            ],
        "End-Point units":  [item["result_unit"] if "result_unit" in item else "" for item in bp_json.get("question3", [])],
        "Raw data metrics": [item["raw_endpoint"] if "raw_endpoint" in item else "raw_endpoint_not_specified" for item in bp_json.get("raw_data_report",[])],
        "Raw data units": [item.get("raw_unit", "") for item in bp_json.get(
            "raw_data_report", [])],
        "SOP(s) for test": bp_json.get("EXPERIMENT", ""),
        "Path/link to sop/protocol": bp_json.get("EXPERIMENT_PROTOCOL", ""),
        "Test start date": bp_json.get("provenance_startdate", datetime.now()),
        "Test end date": bp_json.get("provenance_enddate", datetime.now()),
        }
    return _header


def get_materials_metadata(json_blueprint):
    sample_group_dict = {}
    for item in json_blueprint.get("METADATA_SAMPLE_INFO"):
        group = item.get("param_sample_group","DEFAULT")
        name = item["param_sample_name"]
        sample_group_dict.setdefault(group, []).append(name)    
    _header = {
        "Select item from Project Materials list": 
        sample_group_dict.get("ID", ["ID"])[0],
        "Material Name": sample_group_dict.get("NAME", ["NAME"])[0],
        "Core chemistry": sample_group_dict.get("CHEMISTRY", ["CHEMISTRY"])[0],
        "CAS No": sample_group_dict.get("CASRN", ["CAS_RN"])[0],
        "Material Supplier": 
        sample_group_dict.get("SUPPLIER", ["SUPPLIER"])[0],
        "Material State": "",
        "Batch": sample_group_dict.get("BATCH", ["BATCH"])[0],
        "Date of preparation": datetime.now()
    }
    return _header


def get_materials_columns(nanomaterial=True):
    if nanomaterial:
        return ["", "ERM identifier", "ID", "Name", "CAS", "type", "Supplier",
                "Supplier code", "Batch", "Core", "BET surface in m²/g"]
    else:
        return ["", "Material identifier", "ID", "Name", "CAS", "type",
                "Supplier", "Supplier code", "Batch", "Core"]


def get_treatment(json_blueprint):
    _maxfields = 15
    tmp = []
    condition_type = None
    for item in json_blueprint.get("conditions",[]):
        name = "conditon_name"
        isreplicate = item["condition_type"].startswith("c_replicate")
        isconcentration = item["condition_type"].startswith("c_concentration")
        if not isreplicate:
            tmp.append({'param_name': "TREATMENT {}".format(item[name].upper()),
                         'type': 'group', 'position': '0', 'position_label': 0,
                         'datamodel': item['condition_type'], "value": ""})
        else:
            if condition_type != isreplicate:
                tmp.append({'param_name': "CONTROLS", 'type': 'group',
                            'position': '0', 'position_label': 0,
                            'datamodel': "c_replicate",
                            "value" : ""})
                tmp.append({'param_name': "Positive controls abbreviations", 
                            'type': 'names', 'position': '0', 'position_label': 0, 
                            'datamodel': "CONTROL", "value": ""})
                tmp.append({'param_name': "Positive controls description",
                            'type': 'names', 'position': '0', 'position_label': 0,
                            'datamodel': "CONTROL", "value": ""})
                tmp.append({'param_name': "Negative controls abbreviations",
                            'type': 'names', 'position': '0', 'position_label': 0,
                            'datamodel': "CONTROL", "value": ""})
                tmp.append({'param_name': "Negative controls description", 
                            'type': 'names', 'position': '0',
                            'position_label': 0, 'datamodel': "CONTROL",
                             "value": ""})
                tmp.append({'param_name': "REPLICATES", 'type': 'group',
                            'position': '0', 'position_label': 0,
                            'datamodel': "c_replicate", "value": ""})
        if "condition_unit" in item:
            tmp.append({'param_name': "{} series unit".format(item[name]),
                        'type': 'names', 'position': '0', 'position_label': 0,
                        'datamodel': item['condition_type'], "value": item["condition_unit"]})
        if not isreplicate:
            tag = item['condition_type'].split('_')[1][0].upper()
            _start = 0 if isconcentration else 1
            tmp.append({'param_name': "{} series labels".format(item[name]), 
                        'type': 'names', 'position': '0', 'position_label': 0,
                        'datamodel': item['condition_type'],
                        "value": [f"{tag}{i}" if i <= 3 else "" for i in range(1, _maxfields + 1)]})
        else:
            _start = 0
        tmp.append({'param_name': "{}".format(item[name]), 'type': 'names',
                    'position': '0', 'position_label': 0, 'datamodel': item['condition_type'], 
                    "value":  [i if i<=(2+_start) else "" for i in range(_start, _maxfields + _start + 1)]})
        if isconcentration:
            tmp.append({'param_name': "Treatment type series", 'type': 'names',
                        'position': '0', 'position_label': 0,
                        'datamodel': "c_treatment", "value": ""})
        condition_type = isreplicate
    return pd.DataFrame(tmp)


def get_nmparser_config(json_blueprint):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    json_file_path = os.path.join(current_directory, 
                                  "../../resource/nmparser/DEFAULT_TABLE.json")
    config = {}
    with open(json_file_path, 'r') as json_file:
        # Load the JSON data from the file
        config = json.load(json_file)
    return config


def create_nested_headers_dataframe(dicts,
                                    keys={"METADATA_PARAMETERS": {'group': 'param_group', 'name': 'param_name', 'unit': 'param_unit'}},
                                    levels=['group', 'name', 'unit'],
                                    lookup={'METADATA_SAMPLE_INFO': "Sample", "METADATA_SAMPLE_PREP": "Sample preparation",
                                        "OTHER_SAMPLEPREP": "",
                                        "raw_data_report": "Raw data", "question3": "Results"},
                                    condition_field=[ "raw_conditions","results_conditions"]
                                    ):
    # Initialize an empty DataFrame
    df = pd.DataFrame()
    # Build global condition metadata lookup
    condition_meta = {
        cond.get("conditon_name"): {
            "unit": cond.get("condition_unit", ""),
            "type": cond.get("condition_type", "")
        }
        for cond in dicts.get("conditions", [])
        if cond.get("conditon_name")
    }    

    # Iterate through the dictionaries
    key_conditions = set()    
    for key in keys:
        params = dicts.get(key, [])
        # Collect all unique conditions for this key
        for param in params:
            for cf in condition_field:
                if cf in param:
                    key_conditions.update(param.get(cf, []))

    # Add one column per unique condition (once per key)
    for cond_name in key_conditions:
        cond_info = condition_meta.get(cond_name, {})
        cond_tags = ["Experimental factors"]
        cond_tags.append(cond_name)
        cond_tags.append("")
        cond_tags.append(cond_info.get("unit",""))
        df[tuple(cond_tags)] = None

    for key in keys:
        params = dicts.get(key, [])
        top_label = lookup.get(key, key)
        for param in params:
            try:
                tags = [top_label]
                for level in levels:
                    _tmp = param.get(keys[key][level], "")
                    tags.append(lookup.get(_tmp, _tmp) )
                df[tuple(tags)] = None
            except Exception as err:
                print(f"Error processing param: {e}")
                continue


    # Create MultiIndex DataFrame
    names = ['']
    names.extend(levels)
    df.columns = pd.MultiIndex.from_tuples(df.columns, names=names)
    return df


def autofit_columns(sheet, cols=None):
    # Autofit column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust for padding and scaling
        sheet.column_dimensions[column_letter].width = adjusted_width    
        # Apply colors to top-level keys
    top_level_colors = {'METADATA_PARAMETERS': 'BDD7EE',
                        'Sample': 'FCE4D6', 'Sample preparation': 'BDD7EE',
                        'Raw data': 'FCE4D6', 'Results': 'BDD7EE'}

    for col_num, value in enumerate(cols):
        top_level = value[0]
        if top_level in top_level_colors:
            clr = top_level_colors.get(top_level, "white")
            pf = PatternFill(start_color=clr, end_color=clr, fill_type="solid")
            sheet.cell(row=1, column=col_num+1).fill = pf

        
def autofit_multilevel(df, worksheet):
    for idx, col in enumerate(df.columns):
        # Find the maximum length of the column header (using the last level of multi-index)
        max_length = max(len(str(level)) for level in col) + 1
        # Set the column width based on the length of the column header
        worksheet.set_column(idx, idx, max_length)


def pchem_format_2excel(file_path_xlsx, json_blueprint):
    _SHEET_INFO = "Provider_informations"
    _SHEET_RAW = "Raw_data_TABLE"
    _SHEET_RESULT = "Results_TABLE"
    _SHEET_MATERIAL = "Materials"
    _SHEET_MEASUREMENT = "Experimental_setup"
    current_script_directory = os.path.dirname(os.path.abspath(__file__))
    #resource_file = os.path.join(current_script_directory, "../../resource/nmparser","template_pchem.xlsx")
    #shutil.copy2(resource_file, file_path_xlsx)
    with pd.ExcelWriter(file_path_xlsx, engine='xlsxwriter', mode='w') as writer:
        # sheet = writer.book["Provider_informations"]
        worksheet = writer.book.add_worksheet(_SHEET_INFO)
        bold_format = writer.book.add_format({'bold': True})
        orange_bg_format = writer.book.add_format({'bg_color': '#FFF2CC'})
        material_format = writer.book.add_format({'bg_color': '#00B0F0'})
        position_format = writer.book.add_format({'bg_color': '#FFC000'})
        #_colors = { "top" : '#002060' , "orange" : '#FFC000', 'skyblue' : '#00B0F0' , 'grey' : '#C0C0C0', 'input' : '#FFF2CC'}
        for t in [("A2", "General information"),
                  ("A10", "Template name"),("A11", "Template version "), ("A12", "Template authors"),
                  ("A13", "Template acknowledgment"), ("A6", "Project"), ("A7", "Workpackage"),
                  ("A8", "Partner"), ("D7", "Study"), ("A15", "Template downloaded"),
                  ("B15", datetime.now().strftime("%Y-%m-%d"))
                 ]:
            worksheet.write(t[0], t[1], bold_format)                
        for t in [("E7", "METHOD"),("C2", "EXPERIMENT"), ("B10", "template_name"),
                  ("B11", "template_status"), ("B12", "template_author"),
                  ("B13", "template_acknowledgment"), ("B6", "provenance_project"),
                  ("B7", "provenance_workpackage"), ("B8","provenance_provider")]:
            worksheet.write(t[0], json_blueprint.get(t[1], ""), orange_bg_format)

        worksheet = writer.book.add_worksheet(_SHEET_RESULT)
        df = create_nested_headers_dataframe(json_blueprint,
                                             keys={"raw_data_report": {
                                                 'name': 'raw_endpoint',
                                                 'type': 'raw_aggregate',
                                                 'unit': 'raw_unit'},
                                                 "question3": {
                                                     'name': 'result_name',
                                                     'type': 'result_aggregate',
                                                     'unit': 'result_unit'}},
                                             levels=['name', 'type', 'unit'],
                                             lookup={
                                                 "raw_data_report": "Raw data",
                                                 "question3": "Results"},
                                             condition_field=[ "raw_conditions","results_conditions"]
                                            )
        #df.insert(0, 'Material ID',None)
        df.insert(0, 'Position_ID',None)
        df.to_excel(writer, sheet_name=_SHEET_RESULT) 
        worksheet = writer.book.get_worksheet_by_name(_SHEET_RESULT)
        worksheet.write('A1', 'Material ID', material_format)
        worksheet.write('A2', ' ', material_format)
        worksheet.write('A3', ' ', material_format)
        worksheet.write('A4', ' ', material_format)      
        worksheet.write('B1', 'Position_ID', position_format)
        worksheet.write('B2', ' ', position_format)
        worksheet.write('B3', ' ', position_format)
        worksheet.write('B4', ' ', position_format)            
        autofit_multilevel(df, worksheet)

        df = create_nested_headers_dataframe(json_blueprint,
                                             keys={"METADATA_PARAMETERS" : {'group' : 'param_group', 'name' : 'param_name', 'unit' : 'param_unit'}})
        
        df.to_excel(writer, sheet_name=_SHEET_MEASUREMENT)
        worksheet = writer.book.get_worksheet_by_name(_SHEET_MEASUREMENT)
        worksheet.write('A1', 'Position_ID',position_format)
        worksheet.write('A2', ' ', position_format)
        worksheet.write('A3', ' ', position_format)
        worksheet.write('A4', ' ', position_format)     
        worksheet.write('A5', 'P1', position_format)     
        worksheet.write('A6', 'P2', position_format)
        position_identifiers_range = "{}!$A5:$A1048576".format(_SHEET_MEASUREMENT)  # Entire column A
        writer.book.define_name('Position_Identifiers', position_identifiers_range)           
        autofit_multilevel(df, worksheet)   
        validation = {
            'validate': 'list',
            'source': '=Position_Identifiers'
        }
        writer.book.get_worksheet_by_name(_SHEET_RESULT).data_validation("B5:B1048576", validation)        

        df = create_nested_headers_dataframe(json_blueprint, keys={
                #"METADATA_SAMPLE_INFO" : {'group' : 'param_sample_group', 'name' : 'param_sample_name'},
                "METADATA_SAMPLE_PREP" : {'group' : 'param_sampleprep_group', 'name' : 'param_sampleprep_name'}},
                levels=['group','name'], 
                lookup={'METADATA_SAMPLE_INFO' : "Sample", "METADATA_SAMPLE_PREP" : "Sample preparation", "group" : ""})
        df.to_excel(writer, sheet_name="SAMPLES")     
        worksheet = writer.book.get_worksheet_by_name("SAMPLES")
        worksheet.write('A1', 'Material ID', material_format)
        worksheet.write('A2', ' ', material_format)
        worksheet.write('A3', ' ', material_format)
        validation = {
            'validate': 'list',
            'source': '=ERM_Identifiers'
        }
        writer.book.get_worksheet_by_name("SAMPLES").data_validation("A4:A1048576", validation)          
        autofit_multilevel(df, worksheet)
        materials_sheet = create_materials_sheet(
            writer.book, writer, materials=_SHEET_MATERIAL, info=None,
            results=[_SHEET_RESULT], material_column="A5:A1048576")
    add_hidden_jsondef(file_path_xlsx, json_blueprint)


def add_hidden_jsondef(file_path_xlsx, json_blueprint):
    try:
        workbook = load_workbook(file_path_xlsx)
        hidden_sheet = workbook.create_sheet("TemplateDesigner")
        hidden_sheet.sheet_state = 'hidden'
        hidden_sheet['A1'] = "uuid"
        hidden_sheet['B1'] = "surveyjs"
        hidden_sheet['A2'] = json_blueprint.get("template_uuid", "")
        hidden_sheet['B2'] = json.dumps(json_blueprint)
        hidden_sheet['A3'] = "version"
        hidden_sheet['B3'] = "1.01"        
        hidden_sheet['B2'].style = NamedStyle(name='hidden', hidden=True)  # Hide the cell
        workbook.save(file_path_xlsx)
    except Exception as err:
        print(err)  


def add_plate_layout(file_path_xlsx, json_blueprint):
    if "data_platelayout" in json_blueprint.get("data_sheets", []):
        platexlsx = "platelayout_{}well.xlsx".format(json_blueprint.get("plate_format", 96) )
        current_script_directory = os.path.dirname(os.path.abspath(__file__))
        resource_file = os.path.join(current_script_directory, "../../resource/nmparser", platexlsx)
        copy_sheets(resource_file, file_path_xlsx)


def copy_sheets(source_file, destination_file):
    # Load the source Excel file
    source_wb = load_workbook(source_file)
    # Load the destination Excel file
    destination_wb = load_workbook(destination_file)
    
    # Iterate over each sheet in the source Excel file
    for sheet_name in source_wb.sheetnames:
        # Get the source sheet
        source_sheet = source_wb[sheet_name]
        # Create a new sheet in the destination file with the same name
        destination_sheet = destination_wb.create_sheet(sheet_name)
        # Iterate over each row in the source sheet
        for row in source_sheet.iter_rows(values_only=True):
            destination_sheet.append(row)
        # Copy formulas from the source sheet to the destination sheet
        for row in source_sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    destination_sheet[cell.coordinate].value = cell.value
                if cell.comment:
                    # Create a new comment on the destination cell
                    destination_sheet[cell.coordinate].comment = Comment(cell.comment.text, cell.comment.author)
                cell_dst = destination_sheet[cell.coordinate]    
                cell_dst.font = copy(cell.font)
                cell_dst.fill = copy(cell.fill)
                cell_dst.border = copy(cell.border)
                cell_dst.alignment = copy(cell.alignment)
                cell_dst.number_format = copy(cell.number_format)
                cell_dst.protection = copy(cell.protection)

    destination_wb.save(destination_file)


def get_template_frame(json_blueprint):
    if "METADATA_SAMPLE_INFO" in json_blueprint:
        df_sample = pd.DataFrame(list(get_materials_metadata(json_blueprint).items()), columns=['param_name', 'value'])
        #df_sample = json2frame(json_blueprint["METADATA_SAMPLE_INFO"],sortby=["param_sample_group"]).rename(columns={'param_sample_name': 'param_name'})
        df_sample["type"] = "names"
        df_sample["position"] = -1
        df_sample["datamodel"] = "METADATA_SAMPLE_INFO"
        df_sample = pd.concat([pd.DataFrame([{'param_name': "Test Material Details", 'type': 'group', 'position' : '0', 'position_label' : 0,'datamodel' : 'METADATA_SAMPLE_INFO', 'value' : ''}],
                                            columns=df_sample.columns), df_sample], ignore_index=True)
    else:
        raise Exception("Missing METADATA_SAMPLE_INFO")

    if "METADATA_SAMPLE_PREP" in json_blueprint:
        df_sample_prep = json2frame(json_blueprint["METADATA_SAMPLE_PREP"],sortby=["param_sampleprep_group"]).rename(columns={'param_sampleprep_name': 'param_name'})
        result_df_sampleprep = iom_format(df_sample_prep, "param_name", "param_sampleprep_group")
        result_df_sampleprep["datamodel"] = "METADATA_SAMPLE_PREP"
        result_df_sampleprep["value"] = ""
    else:
        raise Exception("Missing METADATA_SAMPLE_PREP")
    if "METADATA_PARAMETERS" in json_blueprint:
        df_params = json2frame(json_blueprint["METADATA_PARAMETERS"], sortby=["param_group"])
        result_df = iom_format(df_params)
        result_df["datamodel"] = "METADATA_PARAMETERS"
        result_df["value"] = ""
    else:
        raise Exception("Missing METADATA_PARAMETERS")

    #print(df_sample.columns,result_df.columns)
    #empty_row = pd.DataFrame({col: [""] * len(result_df.columns) for col in result_df.columns})
    treatment = get_treatment(json_blueprint)

    df_method = pd.DataFrame(list(get_method_metadata(json_blueprint).items()), columns=['param_name', 'value'])
    df_method["type"] = "names"
    df_method["position"] = -1
    df_method["datamodel"] = "METHOD"
    for df in [df_method, df_sample, result_df_sampleprep, result_df, treatment]:
        if not ("value" in df.columns):
            print(df.columns)
    df_info = pd.concat([
        df_method[["param_name", "type", "position", "datamodel", "value"]],
        df_sample[["param_name", "type", "position", "datamodel", "value"]],
        result_df_sampleprep[["param_name", "type", "position", "datamodel", "value"]],
        result_df[["param_name", "type", "position", "datamodel", "value"]],
        treatment[["param_name", "type", "position", "datamodel", "value"]]
        ], ignore_index=True)
    #print(df_info)
#:END: Please do not add information below this line
#Template version	{{ || version }}
#Template authors	{{ || acknowledgements }}
#Template downloaded	{{ || downloaded }}
    df_info["position"] = range(1, 1 + len(df_info) )
    df_info["position_label"] = 0
    df_info = pd.concat([df_info, pd.DataFrame([{ "param_name" : "Linked exeriment identifier", "type" : "names", "position" : 1, "position_label" : 5 , "datamodel" : "INVESTIGATION_UUID","value" : ""}])])
    df_conditions = pd.DataFrame(json_blueprint["conditions"])
    if "data_sheets" not in json_blueprint:
        json_blueprint["data_sheets"] = ["data_processed"]
    if "data_processed" in json_blueprint["data_sheets"]:    
        df_result = pd.DataFrame(json_blueprint["question3"]) if 'question3' in json_blueprint else None
    else:
        df_result = None
    if "data_raw" in json_blueprint["data_sheets"]:
        df_raw = pd.DataFrame(json_blueprint["raw_data_report"]) if "raw_data_report" in json_blueprint else None
    else:
        df_raw = None
    if "data_calibration" in json_blueprint["data_sheets"]:
        df_calibration = pd.DataFrame(json_blueprint["calibration_report"]) if "calibration_report" in json_blueprint else None
    else:
        df_calibration = None    
    return df_info, df_result, df_raw, df_conditions, df_calibration


def get_unit_by_condition_name(json_blueprint, name):
    for condition in json_blueprint['conditions']:
        if condition['condition_name'] == name:
            return condition.get('condition_unit', None)
    return None


def results_table(df_result, df_conditions=None,
                  result_name='result_name',
                  result_unit='result_unit',
                  results_conditions='results_conditions', sample_column="Material"):
    result_names = df_result[result_name]
    try:
        result_unit = df_result[result_unit]
    except Exception as err:
        result_unit = None

    header1 = list([sample_column])
    header2 = list([""])
    if results_conditions in df_result.columns:
        unique_conditions = set(condition for conditions in df_result[results_conditions].dropna() for condition in conditions)
        header1 = header1 + list(unique_conditions)
        for c in list(unique_conditions):
            try:
                unit = df_conditions.loc[df_conditions['conditon_name'] == c, 'condition_unit'].iloc[0]
                header2 = header2 + [unit if not pd.isnull(unit) else ""]
            except Exception:
                header2 = header2 + [""]

    header1 = header1 + list(result_names)
    if result_unit is not None:
        header2 = header2 + list(result_unit)
        return pd.DataFrame([header2], columns=header1)
    else:
        return pd.DataFrame(columns=header1)


def iom_format_2excel(
        file_path, df_info, df_result, 
        df_raw=None, df_conditions=None, df_calibration=None):
    _SHEET_INFO = "Test_conditions"
    _SHEET_RAW = "Raw_data_TABLE"
    _SHEET_RESULT = "Results_TABLE"
    _SHEET_CALIBRATION = "Calibration_TABLE"
    _SHEET_MATERIAL = "Materials"
    _guide = [
    "Please complete all applicable fields below as far as possible. Aim to familiarise yourself with the Introductory Guidance and Example Filled Templates.",
    "While aiming to standardise data recording as far as we can, flexibility may still be needed for some Test/Assay types and their results:",
    "Thus it may be necessary to add additional items e.g. for further replicates, concentrations, timepoints, or other variations on inputs, results outputs, etc.",
    "If so, please highlight changes & alterations e.g. using colour, and/or comments in notes, or adjacent to data/tables to flag items, fluctuations from norm, etc."
    ]
    _colors = { "top": '#002060' , "orange": '#FFC000', 'skyblue': '#00B0F0' , 'grey': '#C0C0C0', 'input': '#FFF2CC'}
    with pd.ExcelWriter(file_path, engine='xlsxwriter', mode='w') as writer:
        startrow = 7
        _sheet = _SHEET_INFO
        workbook = writer.book
        worksheet = workbook.add_worksheet(_sheet)
        worksheet.set_column(1, 1, 20)
        #writer.sheets[_sheet]
        cell_format_def = {
                    "group":  {'bg_color': _colors['grey'], 'font_color': 'blue', 'text_wrap': True, 'bold': True},
                    "names": {'bg_color': _colors['input'], 'text_wrap': True, 'align': 'right'},
                    "group_labels": {'bg_color': _colors['grey'], 'font_color': 'blue', 'text_wrap': True, 'bold': True},
                    "names_labels": { 'align': 'right', 'bold': True},
                    "top1": {'bg_color': _colors["top"], 'font_color': 'white', 'text_wrap': False, 'font_size': 14, 'bold': True},
                    "top7": {'bg_color': _colors["top"], 'font_color': 'white', 'text_wrap': False, 'font_size': 11, 'bold': True},
                    "orange": {'bg_color': _colors["orange"], 'font_color': 'blue', 'text_wrap': False, 'font_size': 12, 'bold': True},
                    "skyblue": {'bg_color': _colors["skyblue"], 'text_wrap': False}
                    }
        cell_format = {}
        for cf in cell_format_def:
            cell_format[cf] = workbook.add_format(cell_format_def[cf])

        for p in df_info['position_label'].unique():
            max_length = df_info.loc[df_info["position_label"] == p]["param_name"].apply(lambda x: len(str(x))).max()
            worksheet.set_column(p, p, max_length + 1)
            worksheet.set_column(p+1, p+1, 20)

        for index, row in df_info.iterrows():
            cf = cell_format[row["type"]]
            cf_labels = cell_format["{}_labels".format(row["type"])]
            worksheet.write(startrow+row['position']-1, row['position_label'], row['param_name'], cf_labels)
            if isinstance(row["value"], datetime):
                vals = [row["value"].strftime("%Y-%m-%d")]
            else:
                vals = row["value"] if isinstance(row["value"], list) else [str(row["value"])]
            for index, value in enumerate(vals):
                worksheet.write(startrow+row['position']-1, row['position_label']+index+1, value, cf)
            if row["type"] == "group":
                worksheet.set_row(startrow+row['position']-1, None, cf_labels)
            else:
                try:
                    worksheet.write_comment(startrow+row['position']-1, row['position_label']+1, row["datamodel"])
                except Exception:
                    #print(row['param_name'],row["datamodel"])
                    pass

        for row in range(1, startrow-2):
            worksheet.set_row(row, None, cell_format["top7"])
            worksheet.write(row, 0, _guide[row-1])

        worksheet.set_row(startrow-2, None, cell_format["orange"])
        worksheet.set_row(startrow-1, None, cell_format["skyblue"])
        worksheet.write("A1", "Project")
        worksheet.write("B1", "Test Data Recording Form (TDRF)")
        worksheet.write("A6", "TEST CONDITIONS")
        worksheet.write("B6", "Please ensure you also complete a Test Method Description Form (TMDF) for this test type")

        #conc_range = "{}!$B$72:$G$72".format(_SHEET_INFO)  # Entire column B
        #workbook.define_name('CONCENTRATIONS', conc_range)
        linksheets = []
        if df_raw is not None:
            _sheet = _SHEET_RAW
            linksheets = [_sheet]
            new_df = results_table(df_raw, df_conditions,
                                    result_name='raw_endpoint',
                                    result_unit= 'raw_unit',
                                    results_conditions='raw_conditions')
            new_df.to_excel(writer, sheet_name=_sheet, index=False, freeze_panes=(2, 0))
            worksheet = writer.sheets[_sheet]
            #print(new_df.columns)
            for i, col in enumerate(new_df.columns):
                worksheet.set_column(i, i, len(col) + 1 )
                if col == "concentration":
                    colname = xl_col_to_name(i)
                    #worksheet.data_validation('{}3:{}1048576'.format(colname,colname), 
                    #                          {'validate': 'list',
                    #                      'source': '=CONCENTRATIONS'})
    
            #worksheet.add_table(3, 1, 1048576, len(new_df.columns), {'header_row': True, 'name': _SHEET_RAW})

        if df_result is not None:
            _sheet = _SHEET_RESULT 
            new_df = results_table(df_result, result_name='result_name', 
                                   results_conditions='results_conditions')
            new_df.to_excel(writer, sheet_name=_sheet, index=False, 
                            freeze_panes=(2, 0))
            worksheet = writer.sheets[_sheet]
            for i, col in enumerate(new_df.columns):
                worksheet.set_column(i, i, len(col) + 1 )
            linksheets.append(_sheet)

        if df_calibration is not None:
            _sheet = _SHEET_CALIBRATION 
            new_df = results_table(df_calibration, result_name='calibration_entry', 
                                   result_unit="calibration_unit",
                                   results_conditions='calibration_conditions',
                                   sample_column="Sample")
            new_df.to_excel(writer, sheet_name=_sheet, index=False, 
                            freeze_panes=(2, 0))
            worksheet = writer.sheets[_sheet]
            for i, col in enumerate(new_df.columns):
                worksheet.set_column(i, i, len(col) + 1 )
            linksheets.append(_sheet)            

        materials_sheet = create_materials_sheet(
            workbook, writer, materials=_SHEET_MATERIAL,
            info=_SHEET_INFO, results=linksheets)


def create_materials_sheet(workbook, writer, materials, info=None, results=[], material_column="A3:A1048576"):
    info_sheet = None if info is None else writer.sheets[info]
    materials_sheet = workbook.add_worksheet(materials)
    column_headers = get_materials_columns()
    table = pd.DataFrame(columns=column_headers)
    table.to_excel(writer, sheet_name=materials, startrow=0, startcol=0, index=False)
    erm_identifiers_range = "{}!$B:$B".format(materials)  # Entire column B
    workbook.define_name('ERM_Identifiers', erm_identifiers_range)
    validation_cell = 'B25'  # cell to apply validation
    validation = {
        'validate': 'list',
        'source': '=ERM_Identifiers'
    }
    if info_sheet is not None:
        info_sheet.data_validation(validation_cell, validation)
        vlookup = [('B26',3),('B27',9),('B28',4),('B29',6),('B31',8)]
        for v in vlookup:
            formula = '=VLOOKUP(B$25,Materials!$B:$J,"{}",FALSE)'.format(v[1])
            info_sheet.write_formula(v[0], formula)
        readonly_format = workbook.add_format({'locked': True})
    for result in results:
        try:
            result_sheet = writer.sheets[result]
            result_sheet.data_validation(material_column, validation)
            #protect_headers(result_sheet,readonly_format)
        except Exception as err:
            print(err)
            pass
    return materials_sheet


def protect_headers(worksheet, readonly_format):
    worksheet.set_default_row(options={'locked': False})
    worksheet.set_row(0, None, readonly_format)
    worksheet.set_row(1, None, readonly_format)
    worksheet.protect()
