#PowerDB VERSION 2.2.5.4
#Created solely by WeDu, published in 6/21/25
import re
import stat
import os
import sys
import openpyxl
class InnerFunctions:
    def __init__(self):
        pass
    @staticmethod
    def count_occurrences(word, string):
        if not isinstance(word, str) or not isinstance(string, str):
            raise TypeError("Inputs must be strings.")
        if not word:
            raise ValueError("word cannot be empty.")
        count = 0
        word_len = len(word)
        string_len = len(string)
        if word_len > string_len:
            return 0
        for i in range(string_len - word_len + 1):
            if string[i:i + word_len] == word:
                count += 1
        return count
    @staticmethod
    def add_data_to_inner_lists(main_list, second_list):
        if not isinstance(main_list, list) or not isinstance(second_list, list):
            raise TypeError("Inputs must be lists.")
        result = []
        for i, inner_item in enumerate(main_list):
            if not isinstance(inner_item, list):
                raise ValueError("Inner items of main_list must be lists.")
            if i < len(second_list):
                result.append(inner_item + [second_list[i]])
            else:
                result.append(inner_item + [None])
                print("Warning: second_list is shorter than expected. Filling with None.", file=sys.stderr)
        return result
    @staticmethod
    def combine_lists(input_list):
        if not isinstance(input_list, list):
            raise TypeError("Input must be a list.")
        output_list = []
        for inner_list in input_list:
            if not isinstance(inner_list, list):
                raise ValueError("Inner items must be lists.")
            output_list.extend(inner_list)
        return output_list
inner_functions = InnerFunctions()
class SharedFunctions:
    def __init__(self):
        pass
    @staticmethod
    def normalize_path(file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = os.path.abspath(os.path.normpath(file_path))
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File '{filepath}' does not exist.")
        return filepath
    def check_file_permissions(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = self.normalize_path(file_path)
        read_write_mode = stat.S_IRUSR | stat.S_IWUSR
        try:
            mode = stat.S_IMODE(os.stat(filepath).st_mode)
            return bool(mode & read_write_mode)
        except OSError as e:
            raise OSError(f"OS Error while checking permissions for '{filepath}': {e}")
    def read_file_bytes(self, file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            with open(filepath, 'rb') as f:
                return f.read()
        except OSError as e:
            raise OSError(f"OS Error while reading file '{filepath}': {e}")
    def write_file_bytes(self, file_path, data):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            with open(filepath, 'wb') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while writing file '{filepath}': {e}")
    def append_file_bytes(self, file_path, data):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = self.normalize_path(filepath)
        if not self.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            with open(filepath, 'ab') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while appending file '{filepath}': {e}")
shared_functions = SharedFunctions()
class CreateOperations:
    def __init__(self):
        pass
    @staticmethod
    def make_db(file_path):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = os.path.abspath(os.path.normpath(filepath))
        if not os.path.exists(filepath):
            c = open(filepath,'x')
            c.close()
            shared_functions.write_file_bytes(filepath,b'#POWER_DB') # <- THE FIX IN 2.2.5.4 UPDATE,
            """ the problem was that if file does exist instead of being ignored the entire file would be overwritten and all data
            in it would be removed and the entire file content would be '#POWER_DB', and now after the correction ONLY when the
            file doesn't exist the file would be created and then the tag would be inserted in it"""
    @staticmethod
    def make_container(file_path, name):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not name:
            raise ValueError("name cannot be empty.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            # Normalize line endings
            file_content_str = file_content_str.replace('\r\n', '\n')
            num = inner_functions.count_occurrences('$<', file_content_str)
            container_string = f"{os.linesep if not file_content_str.endswith('\n') else ''}$<{num},{name}>"
            container_bytes = container_string.encode('utf-8')
            if f'$<{num},{name}>' not in file_content_str:
                shared_functions.append_file_bytes(filepath, container_bytes)
            else:
                print(f"Container '{name}' already exists.")
        except OSError as e:
            print(f"OS Error: {e}")
    @staticmethod
    def make_table(file_path, name):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not name:
            raise ValueError("name cannot be empty.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')
            num = inner_functions.count_occurrences('&<', file_content_str)
            table_string = f"{os.linesep if not file_content_str.endswith('\n') else ''}&<{num}^{name}>"
            table_bytes = table_string.encode('utf-8')
            if f'&<{num}^{name}>' not in file_content_str:
                shared_functions.append_file_bytes(filepath, table_bytes)
            else:
                print(f"Table '{name}' already exists.")
        except OSError as e:
            print(f"OS Error: {e}")
create = CreateOperations()
class container_data_class:
    def __init__(self):
        pass
    @staticmethod
    def getname(file, ID):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(ID, int):
            raise TypeError("id must be an int")
        if ID < 0:
            raise ValueError("id must be >= 0")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            pattern = rf'\$<({ID}),([^>]*)>'
            match = re.search(pattern, r)
            if match:
                return match.group(2)
            else:
                return None
        except OSError:
            return None
    @staticmethod
    def getid(file, name, plogic=True):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        if not name:
            raise ValueError("name cannot be empty")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            pattern = rf'\$<(\d+),{re.escape(name)}>'
            match = re.search(pattern, r)
            if match:
                return int(match.group(1)) if plogic else int(match.group(1)) + 1
            else:
                return -1 if plogic else 0
        except OSError:
            return -1 if plogic else 0
    def insert(self, file, data, address=None, showrelational=False):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(data, str):
            raise TypeError("data must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or none")
        if not isinstance(showrelational, bool):
            raise TypeError("showrelational must be a bool")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >= 0")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            num_sectors = self.numbersectors(filepath, containerid)
            if showrelational:
                print(sectorid, num_sectors)
            if not self._check(filepath, 'sector', [containerid, sectorid]):
                if sectorid - num_sectors <= 1:
                    new_data = f"!<[{containerid},{sectorid}],{data}>!".encode('utf-8', errors='surrogateescape')
                    if r.endswith('\n'): # Changed from os.linesep
                        shared_functions.append_file_bytes(filepath, new_data)
                    else:
                        shared_functions.append_file_bytes(filepath, '\n'.encode('utf-8', errors='surrogateescape') + new_data) # changed from os.linesep
        except OSError as e:
            print(f"OS Error: {e}")
    @staticmethod
    def read(file, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or none")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >= 0")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            pattern = rf'(?s)!<\[{containerid},{sectorid}],(.*?)>!'
            match = re.search(pattern, r)
            if match:
                return match.group(1)
            else:
                return ""
        except OSError as e:
            print(f"OS Error: {e}")
            return ""
    def edit(self, file, data, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(data, str):
            raise TypeError("data must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or None")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers.")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >=0.")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            if self._check(filepath, 'sector', [containerid, sectorid]):
                pattern = rf'(!<\[{containerid},{sectorid}],)([^>!]*)>!'
                replacement = rf'!<[{containerid},{sectorid}],{data}>!'
                new_content = re.sub(pattern, replacement, r, count=1).encode('utf-8', errors='surrogateescape')
                shared_functions.write_file_bytes(filepath, new_content)
        except OSError as e:
            print(f"OS Error: {e}")
    @staticmethod
    def change_name(file, new_name, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(new_name, str):
            raise TypeError("new_name  must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("Container ID must be >= 0")
        if not new_name:
            raise ValueError("new_name cannot be empty")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n','\n') # Normalize line endings
            pattern = rf"\$<{containerid},([^>]*)>"
            replacement = f"$<{containerid},{new_name}>\n"
            updated_content = re.sub(pattern, replacement, file_content_str).encode('utf-8', errors='surrogateescape')
            lines = updated_content.splitlines()
            non_empty_lines = [line for line in lines if line.strip()]
            final_content = '\n'.encode('utf-8', errors='surrogateescape').join(non_empty_lines) # changed from os.linesep
            shared_functions.write_file_bytes(filepath, final_content)
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    @staticmethod
    def readsectors(file, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            r = r.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf'(?s)!<\[{containerid},(\d+)],(.*?)>!'
            matches = re.finditer(pattern, r)
            section_data = []
            for match in matches:
                section_id = int(match.group(1))
                value = match.group(2)
                section_data.append((section_id, value))
            section_data.sort(key=lambda x: x[0])
            data = [value for _, value in section_data]
            return data
        except OSError as e:
            print(f"OS Error: {e}")
            return []
    @staticmethod
    def numbercontainers(file, plogic=False):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            r = r.replace('\r\n', '\n')  # Normalize line endings
            count = len(re.findall(r'\$<(\d+),', r))
            return count if not plogic else count - 1
        except OSError as e:
            print(f"OS Error: {e}")
            return -1 if plogic else 0
    @staticmethod
    def numbersectors(file, containerid, plogic=False):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
            r = r.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf'!<\[{containerid},\d+]'
            count = len(re.findall(pattern, r))
            return count if not plogic else count - 1
        except OSError as e:
            print(f"OS Error: {e}")
            return -1 if plogic else 0
    @staticmethod
    def delete(file, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or none")
        if address is None:
            address = []
        if len(address) != 2:
            raise ValueError("address must contain containerid and sectorid.")
        containerid, sectorid = address
        if not isinstance(containerid, int) or not isinstance(sectorid, int):
            raise TypeError("containerid and sectorid must be integers.")
        if containerid < 0 or sectorid < 0:
            raise ValueError("containerid and sectorid must be >= 0.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # normalize line endings.
            pattern = rf'(?s)!<\[{containerid},{sectorid}],(.*?)>!'
            updated_content = re.sub(pattern, '', file_content_str)
            lines = updated_content.split('\n')  # changed from os.linesep
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            shared_functions.write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            print(f"OS Error: {e}")
    @staticmethod
    def drop(file, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')
            container_name_pattern = rf"\$<{containerid},[^>]*>\n?"
            updated_content = re.sub(container_name_pattern, '', file_content_str)
            sector_pattern = rf"(?s)!<\[{containerid},\d+],(?:(?!>!).)*?>!\n?"
            updated_content = re.sub(sector_pattern, '', updated_content)
            lines = updated_content.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)
            if final_content and not final_content.endswith('\n'):
                final_content += '\n'
            shared_functions.write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            print(f"OS Error: {e}")
    @staticmethod
    def _check(file, type_, address):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(type_, str):
            raise TypeError("type must be a string.")
        if not isinstance(address, list):
            raise TypeError("address must be a list.")
        if type_ not in ('container', 'sector'):
            raise ValueError("type must be 'container' or 'sector'")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if type_ == 'container':
            if len(address) != 2:
                raise ValueError("address for container check must contain containerid and name")
            containerid, name = address
            if not isinstance(containerid, int):
                raise TypeError("containerid must be an integer.")
            if not isinstance(name, str):
                raise TypeError("name must be a string.")
            if containerid < 0:
                raise ValueError("containerid must be >= 0")
            try:
                r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
                r = r.replace('\r\n', '\n')  # Normalize line endings
                pattern = rf'\$<{containerid},{re.escape(name)}>'
                return bool(re.search(pattern, r))
            except FileNotFoundError:
                return False
            except OSError:
                return False
        elif type_ == 'sector':
            if len(address) != 2:
                raise ValueError("address for sector check must contain containerid and sectorid")
            containerid, sectorid = address
            if not isinstance(containerid, int) or not isinstance(sectorid, int):
                raise TypeError("containerid and sectorid must be integers.")
            if containerid < 0 or sectorid < 0:
                raise ValueError("containerid and sectorid must be >= 0.")
            try:
                r = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
                r = r.replace('\r\n', '\n')  # Normalize line endings
                pattern = rf'!<\[{containerid},{sectorid}]'
                return bool(re.search(pattern, r))
            except FileNotFoundError:
                return False
            except OSError:
                return False
        return False
    def reindex_sections(self, file, containerid):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(containerid, int):
            raise TypeError("containerid must be an integer.")
        if containerid < 0:
            raise ValueError("containerid must be >= 0")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            container_name = self.getname(filepath, containerid)
            if container_name is None:
                print(f"Error: Container with ID {containerid} not found. Cannot reindex.")
                return
            original_data_values = self.readsectors(filepath, containerid)
            print(f"DEBUG: original_data_values content: {original_data_values}")
            self.drop(filepath, containerid)
            try:
                new_container_line = f"$<{containerid},{container_name}>\n"
                current_content = shared_functions.read_file_bytes(filepath).decode('utf-8', errors='surrogateescape')
                if current_content and not current_content.endswith('\n'):
                    current_content += '\n'
                print(new_container_line)
                shared_functions.append_file_bytes(filepath,
                                                   new_container_line.encode('utf-8', errors='surrogateescape'))
            except OSError as e:
                print(f"OS Error inserting container definition: {e}")
                return
            new_sector_id = 0
            for data_value in original_data_values:
                self.insert(filepath, data_value, address=[containerid, new_sector_id])
                new_sector_id += 1
            print(f"Sections for container {containerid} reindexed successfully.")
        except ValueError as e:
            print(f"Error during reindexing: Data format issue. {e}. Please check the content returned by readsectors.")
        except OSError as e:
            print(f"OS Error during reindexing: {e}")
        except Exception as e:
            print(f"An unexpected error occurred during reindexing: {e}")
container_data = container_data_class()
class table_data_class:
    def __init__(self):
        super().__init__()
    def _validate_dbfile(self, dbfile):
        if not isinstance(dbfile, str):
            raise TypeError("dbfile must be a string.")
        dbfile_path = self.normalize_path(dbfile)
        return dbfile_path
    @staticmethod
    def normalize_path(file_path, C: bool = False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        filepath = os.path.abspath(os.path.normpath(file_path))
        if not os.path.exists(filepath):
            if C:
                return filepath
            else:
                raise FileNotFoundError(f"File '{filepath}' does not exist.")
        return filepath
    def append_file_bytes(self, file_path, data):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes.")
        filepath = file_path
        if not file_path.lower().endswith('.pdb'):
            filepath = f'{file_path}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            with open(filepath, 'ab') as f:
                f.write(data)
        except OSError as e:
            raise OSError(f"OS Error while appending file '{filepath}': {e}")
    def getname(self, file_path, table_id):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(table_id, int):
            raise TypeError("table_id must be an integer.")
        if table_id < 0:
            raise ValueError("table_id must be >= 0.")
        filepath = file_path
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings
            pattern = rf'&<{table_id}\^([^>]*)>'
            match = re.search(pattern, file_content_str)
            if match:
                return match.group(1)
            return None
        except OSError:
            return None
    def getid(self, file_path, name, plogic=True):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(name, str):
            raise TypeError("name must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        if not name:
            raise ValueError("name cannot be empty")
        filepath = file_path
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings
            pattern = rf'&<(\d+)\^{re.escape(name)}>'
            match = re.search(pattern, file_content_str)
            if match:
                return int(match.group(1)) if plogic else int(match.group(1)) + 1
            return -1 if plogic else 0
        except OSError:
            return -1 if plogic else 0
    def numbertables(self, file_path, plogic=False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = file_path
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings
            count = len(re.findall(r'&<\d+\^', file_content_str))
            return count if not plogic else count - 1
        except OSError:
            return -1 if plogic else 0
    def hcolumn(self, file: str, tableid: int, plogic: bool = False, sprow: int = -1):
        try:
            rawp = file
            if not rawp.lower().endswith('.pdb'):
                rawp = f'{rawp}.pdb'
            rawp = self.normalize_path(rawp)
            if not shared_functions.check_file_permissions(rawp):
                raise PermissionError(f"Insufficient permissions for I/O to '{rawp}'")
            f = open(rawp, 'rb')
            rawd = f.read()
            f.close()
            r = rawd.decode('utf-8', errors='surrogateescape')
        except (FileNotFoundError,OSError,PermissionError):
            return -1 if plogic else 0
        if sprow == -1:
            pattern = rf'~<\[{tableid};(\d+)\?'
        else:
            pattern = rf'~<\[{tableid};(\d+)\?{sprow}\]'
        matches = re.findall(pattern, r)
        if not matches:
            return -1 if plogic else 0
        try:
            max_col = max(map(int, matches))
            return max_col if plogic else max_col + 1
        except ValueError:
            return -1 if plogic else 0
    def hrow(self, file: str, tableid: int, plogic: bool = False, sprow: int = -1):
        try:
            rawp = file
            if not rawp.lower().endswith('.pdb'):
                rawp = f'{rawp}.pdb'
            rawp = self.normalize_path(rawp)
            if not shared_functions.check_file_permissions(rawp):
                raise PermissionError(f"Insufficient permissions for I/O to '{rawp}'")
            f = open(rawp, 'rb')
            rawd = f.read()
            f.close()
            r = rawd.decode('utf-8', errors='surrogateescape')
        except (FileNotFoundError, OSError, PermissionError):
            return -1 if plogic else 0
        if sprow == -1:
            pattern = rf'~<\[{tableid};\d+\?(\d+)]'
        else:
            pattern = rf'~<\[{tableid};{sprow}\?(\d+)]' #corrected line.
        matches = re.findall(pattern, r)
        if not matches:
            return -1 if plogic else 0
        try:
            max_row = max(map(int, matches))
            return max_row if plogic else max_row + 1
        except ValueError:
            return -1 if plogic else 0
    def numbercolumns(self, file_path, address=None, plogic=False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or None.")
        if address is not None:
            if len(address) != 2:
                raise ValueError("address list must have length 2: (table_id, row_id)")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = file_path
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            if address:
                return self.hcolumn(filepath, address[0], plogic, address[1])
            return -1 if plogic else 0
        except OSError:
            return -1 if plogic else 0
    def numberrows(self, file_path, address=None, plogic=False):
        if not isinstance(file_path, str):
            raise TypeError("file_path must be a string.")
        if address is not None and not isinstance(address, list):
            raise TypeError("address must be a list or None.")
        if address is not None:
            if len(address) != 2:
                raise ValueError("address list must have length 2: [table_id, row_id]")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = file_path
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            if address:
                return self.hrow(filepath, address[0], plogic, address[1])
            return -1 if plogic else 0
        except OSError:
            return -1 if plogic else 0
    def totalcolumns(self, file: str, tableid: int, plogic: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            return self.hcolumn(filepath, tableid, plogic)
        except OSError:
            return -1 if plogic else 0
    def totalrows(self, file: str, tableid: int, plogic: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            return self.hrow(filepath, tableid, plogic)
        except OSError:
            return -1 if plogic else 0
    def totaltable(self, file: str, tableid: int, plogic: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(plogic, bool):
            raise TypeError("plogic must be a bool.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            return [self.hcolumn(filepath, tableid, plogic), self.hrow(filepath, tableid, plogic)]
        except OSError:
            return [-1 if plogic else 0, -1 if plogic else 0]
    def insert(self, file: str, data: str, address=None, showmatrix: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(data, str):
            raise TypeError("data must be a string")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            info = self.totaltable(filepath, tableid)
            if showmatrix:
                print(columnid, info[0])
                print(rowid, info[1])
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            if not re.search(pattern, r):
                if columnid <= info[0] and rowid <= info[1]:
                    self.append_file_bytes(filepath,
                                            f"{'\n' if r and not r.endswith('\n') else ''}~<[{tableid};{columnid}?{rowid}],{data}>~".encode(
                                                'utf-8', errors='surrogateescape'))  # changed os.linesep
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def read(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            match = re.search(pattern, r)
            if match:
                return match.group(1)
            return ""
        except OSError:
            return ""
        except TypeError:
            return ""
        except ValueError:
            return ""
        except IndexError:
            return ""
    def readcolumns(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 2:
            raise ValueError("address must contain tableid and rowid")
        tableid = address[0]
        rowid = address[1]
        if not isinstance(tableid, int) or not isinstance(rowid, int):
            raise TypeError("tableid and rowid must be integers")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            pattern = rf"(?s)~<\[{tableid};(\d+)\?{rowid}],(.*?)>~"
            matches = re.finditer(pattern, r)
            column_data = []
            for match in matches:
                column_id = int(match.group(1))
                value = match.group(2)
                column_data.append((column_id, value))
            column_data.sort(key=lambda x: x[0])
            data = [value for _, value in column_data]
            return data
        except OSError:
            return []
        except TypeError:
            return []
        except ValueError:
            return []
        except IndexError:
            return []
    def readrows(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 2:
            raise ValueError("address must contain tableid and columnid")
        tableid = address[0]
        columnid = address[1]
        if not isinstance(tableid, int) or not isinstance(columnid, int):
            raise TypeError("tableid and columnid must be integers")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            r = file_content_str
            row_data = []
            pattern = rf"(?s)~<\[{tableid};{columnid}\?(\d+)],(.*?)>~"
            for match in re.finditer(pattern, r):
                row_id = int(match.group(1))
                value = match.group(2)
                row_data.append((row_id, value))
            row_data.sort(key=lambda x: x[0])
            data = [value for _, value in row_data]
            return data
        except OSError:
            return []
        except TypeError:
            return []
        except ValueError:
            return []
        except IndexError:
            return []
    def edit(self, file: str, data: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(data, str):
            raise TypeError("data must be a string")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            replacement = rf"~<[{tableid};{columnid}?{rowid}],{data}>~"
            updated_content = re.sub(pattern, replacement, file_content_str)
            shared_functions.write_file_bytes(filepath, updated_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def change_name(self, file: str, new_name: str, tableid: int):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(new_name, str):
            raise TypeError("new_name must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"&<{tableid}\^.*?\n"  # changed from os.linesep
            replacement = f"&<{tableid}^{new_name}>\n"  # changed from os.linesep
            updated_content = re.sub(pattern, replacement, file_content_str)
            lines = updated_content.splitlines()
            non_empty_lines = [line for line in lines if line.strip()]
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            shared_functions.write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def all_addresses_grouping(self, file: str, tableid: int, filtermode: int):
        # filtermode must be either 0(columns) or 1(rows)
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(filtermode, int):
            raise TypeError("filtermode must be an integer")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"~<\[{tableid};(\d+)\?(\d+)]"
            matches = re.findall(pattern, file_content_str)
            addresses = [[tableid, int(col), int(row)] for col, row in matches]
            addresses.sort(key=lambda address_: (address_[1], address_[2]))
            if filtermode < 2:
                grouped = {}
                index = filtermode + 1  # 1 for col, 2 for row
                for address in addresses:
                    key = address[index]
                    if key not in grouped:
                        grouped[key] = []
                    grouped[key].append(address)
                return list(grouped.values())
            else:
                return []
        except OSError:
            return []
        except TypeError:
            return []
        except ValueError:
            return []
        except IndexError:
            return []
    def all_addresses_list(self, file: str, tableid: int, totalnum: bool = False):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        if not isinstance(totalnum, bool):
            raise TypeError("totalnum must be a boolean")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"~<\[{tableid};(\d+)\?(\d+)]"
            matches = re.findall(pattern, file_content_str)
            addresses = [[tableid, int(col), int(row)] for col, row in matches]
            addresses.sort(key=lambda address: (address[1], address[2]))
            return len(addresses) if totalnum else addresses
        except OSError:
            return 0 if totalnum else []
        except TypeError:
            return 0 if totalnum else []
        except ValueError:
            return 0 if totalnum else []
        except IndexError:
            return 0 if totalnum else []
    def delete(self, file: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        if address is None:
            address = []
        if not isinstance(address, list):
            raise TypeError("address must be a list")
        if len(address) != 3:
            raise ValueError("address must contain tableid, columnid, and rowid")
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        if not isinstance(tableid, int) or not isinstance(columnid, int) or not isinstance(rowid, int):
            raise TypeError("tableid, columnid, and rowid must be integers")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            pattern = rf"(?s)~<\[{tableid};{columnid}\?{rowid}],(.*?)>~"
            updated_content = re.sub(pattern, '', file_content_str)
            lines = updated_content.split('\n')  # changed from os.linesep
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)  # changed from os.linesep
            shared_functions.write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def drop(self, file: str, tableid: int):
        if not isinstance(file, str):
            raise TypeError("file must be a string")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n')  # Normalize line endings
            # Corrected Regex for multi-line data
            pattern = rf"&<{tableid}\^.*?\n(?:~<\[{tableid};\d+\?\d+\](?:(?!~<\[).)*?>~(?:\n.*?)*?\n?)*"  # changed from os.linesep
            updated_content = re.sub(pattern, '', file_content_str, flags=re.MULTILINE | re.DOTALL)
            # Remove empty lines that were created by re.sub
            lines = updated_content.split('\n')  # changed from os.linesep
            non_empty_lines = [line for line in lines if line.strip() != '']
            final_content = '\n'.join(non_empty_lines)
            if final_content and not final_content.endswith('\n'):
                final_content += '\n'
            shared_functions.write_file_bytes(filepath, final_content.encode('utf-8', errors='surrogateescape'))
        except OSError as e:
            raise OSError(f"OS Error: {e}")
        except TypeError as e:
            raise TypeError(f"Type Error: {e}")
        except ValueError as e:
            raise ValueError(f"Value Error: {e}")
        except IndexError as e:
            raise IndexError(f"Index Error: {e}")
    def export_tables_to_excel(self, dbfile: str, file_path: str):
        if not isinstance(dbfile, str):
            raise TypeError("filepath must be a string.")
        dbfile_path = dbfile
        if not dbfile_path.lower().endswith('.pdb'):
            dbfile_path = f'{dbfile_path}.pdb'
        dbfile_path = self._validate_dbfile(dbfile_path)
        if not isinstance(file_path, str):
            raise TypeError("filepath must be a string.")
        filepath = file_path
        if not filepath.lower().endswith(('.xlsx', '.xlsm')):
            filepath = f"{filepath}.xlsx"
        filepath = self.normalize_path(filepath, True)
        if not shared_functions.check_file_permissions(dbfile_path):
            raise PermissionError(f"Insufficient permissions for I/O to '{dbfile_path}'")
        try:
            num_sheets = self.numbertables(dbfile_path, False)
            total_items = 0
            for table_id in range(num_sheets):
                total_items += self.all_addresses_list(dbfile_path, table_id, True)
            data_list = []
            raw_data = []
            stuff_list = []
            for main in range(num_sheets):
                raw_data.append(self.all_addresses_list(dbfile_path, main))
            sraw_data = inner_functions.combine_lists(raw_data)
            for m in range(total_items):
                address = sraw_data[m]
                try:
                    value = self.read(dbfile_path, [address[0], address[1], address[2]])
                    stuff_list.append(value)
                except Exception as e:
                    print(f"Error reading data at {address}: {e}", file=sys.stderr)
                    stuff_list.append(None)
            data_list.append(inner_functions.add_data_to_inner_lists(sraw_data, stuff_list))
            data_list = inner_functions.combine_lists(data_list)
            created_sheets = {}
            if os.path.isfile(filepath):
                workbook = openpyxl.load_workbook(filepath)
            else:
                workbook = openpyxl.Workbook()
                std = workbook['Sheet']
                workbook.remove(std)
            for item in range(len(data_list)):
                if len(data_list[item]) == 4:
                    table_id, col_id_0, row_id_0, value = data_list[item]
                    table_name = self.getname(dbfile_path, table_id)
                    if not isinstance(table_name, str) or not re.match(r'^[a-zA-Z0-9_\- ]+$', table_name):
                        print(f"Warning: Invalid table name '{table_name}'. Skipping data entry.")
                        continue
                    if table_name not in created_sheets:
                        if table_name not in workbook.sheetnames:
                            workbook.create_sheet(table_name)
                        created_sheets[table_name] = True
                    if table_name in workbook.sheetnames:
                        sheet = workbook[table_name]
                        col_id_1 = col_id_0 + 1
                        row_id_1 = row_id_0 + 1
                        sheet.cell(row=row_id_1, column=col_id_1, value=value)
                    else:
                        print(f"Warning: Sheet '{table_name}' not found.")
                else:
                    print(f"Warning: Invalid data item {item}. Expected [table_id, column_id, row_id, value].")
            workbook.save(filepath)
            print(f"Data inserted into '{filepath}'.")
        except (TypeError, FileNotFoundError, OSError) as e:
            print(f"Error: {e}", file=sys.stderr)
            raise
        except Exception as e:
            print(f"An unexpected error occurred: {e}", file=sys.stderr)
            raise
    def reindex_table(self, file: str, tableid: int, modeCR: bool):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(tableid, int):
            raise TypeError("tableid must be an integer.")
        if not isinstance(modeCR, bool):
            raise ValueError("mode must be True or False.")
        start_column_index = 0
        start_row_index = 0
        filepath = file
        if not filepath.lower().endswith('.pdb'):
            filepath = f'{filepath}.pdb'
        filepath = self.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            table_name = self.getname(filepath, tableid)
            if table_name is None:
                print(f"Error: Table with ID {tableid} not found. Cannot reindex.")
                return
            all_addresses = self.all_addresses_list(filepath, tableid)
            if not all_addresses:
                print(f"No data found for table {tableid}. Reindexing skipped.")
                self.drop(filepath, tableid)
                current_content_bytes = shared_functions.read_file_bytes(filepath)
                current_content_str = current_content_bytes.decode('utf-8', errors='surrogateescape')
                if not current_content_str.startswith("#POWER_DB"):
                    current_content_str = "#POWER_DB\n" + current_content_str
                current_content_str = re.sub(rf"&<{tableid}\^[^>]*>\n?", "", current_content_str)
                if current_content_str and not current_content_str.endswith('\n'):
                    current_content_str += '\n'
                new_table_line = f"&<{tableid}^{table_name}>\n"
                final_content_str = current_content_str + new_table_line
                shared_functions.write_file_bytes(filepath, final_content_str.encode('utf-8', errors='surrogateescape'))
                return
            collected_cells_data = []
            for addr in all_addresses:
                col_id = addr[1]
                row_id = addr[2]
                data = self.read(filepath, address=[tableid, col_id, row_id])
                collected_cells_data.append({'col': col_id, 'row': row_id, 'data': data})
            self.drop(filepath, tableid)
            current_content_bytes = shared_functions.read_file_bytes(filepath)
            current_content_str = current_content_bytes.decode('utf-8', errors='surrogateescape')
            if not current_content_str.startswith("#POWER_DB"):
                current_content_str = "#POWER_DB\n" + current_content_str
            current_content_str = re.sub(rf"&<{tableid}\^[^>]*>\n?", "", current_content_str)
            if current_content_str and not current_content_str.endswith('\n'):
                current_content_str += '\n'
            new_table_line = f"&<{tableid}^{table_name}>\n"
            final_content_str = current_content_str + new_table_line
            shared_functions.write_file_bytes(filepath, final_content_str.encode('utf-8', errors='surrogateescape'))
            if modeCR:
                sorted_cells = sorted(collected_cells_data, key=lambda x: (x['col'], x['row']))
                unique_cols = sorted(list(set(c['col'] for c in sorted_cells)))
                col_id_map = {old_col: new_idx + start_column_index for new_idx, old_col in enumerate(unique_cols)}
                current_row_for_new_col = {}
                for cell in sorted_cells:
                    original_col = cell['col']
                    data = cell['data']
                    new_col = col_id_map[original_col]
                    if new_col not in current_row_for_new_col:
                        current_row_for_new_col[new_col] = start_row_index
                    new_row = current_row_for_new_col[new_col]
                    self.insert(filepath, data, address=[tableid, new_col, new_row])
                    current_row_for_new_col[new_col] += 1
            else:
                sorted_cells = sorted(collected_cells_data, key=lambda x: (x['row'], x['col']))
                unique_rows = sorted(list(set(c['row'] for c in sorted_cells)))
                row_id_map = {old_row: new_idx + start_row_index for new_idx, old_row in enumerate(unique_rows)}
                current_col_for_new_row = {}
                for cell in sorted_cells:
                    original_row = cell['row']
                    data = cell['data']
                    new_row = row_id_map[original_row]
                    if new_row not in current_col_for_new_row:
                        current_col_for_new_row[new_row] = start_column_index
                    new_col = current_col_for_new_row[new_row]
                    self.insert(filepath, data, address=[tableid, new_col, new_row])
                    current_col_for_new_row[new_row] += 1
            print(f"Table {tableid} reindexed successfully in '{"column to row" if modeCR is True else "row to column"}' mode.")
        except ValueError as e:
            print(f"Error during reindexing: {e}")
        except OSError as e:
            print(f"OS Error during reindexing: {e}")
        except Exception as e:
            print(f"An unexpected error occurred during reindexing: {e}")
table_data = table_data_class()
class OtherClass:
    def __init__(self):
        pass
    @staticmethod
    def clear(file: str):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            shared_functions.write_file_bytes(filepath, b'#POWER_DB')
        except OSError as e:
            raise OSError(f"Error occurred while clearing/writing to file: {e}")
    @staticmethod
    def check(file: str, itemtype: str, address=None):
        if not isinstance(file, str):
            raise TypeError("file must be a string.")
        if not isinstance(itemtype, str):
            raise TypeError("itemtype must be a string.")
        if itemtype.lower() not in ('container', 'table', 'sector', 'cell'):
            raise ValueError("Invalid itemtype.  Must be 'container', 'table', 'sector', or 'cell'.")
        filepath = file
        if not file.lower().endswith('.pdb'):
            filepath = f'{file}.pdb'
        filepath = shared_functions.normalize_path(filepath)
        if not shared_functions.check_file_permissions(filepath):
            raise PermissionError(f"Insufficient permissions to I/O file: '{filepath}'")
        try:
            file_content_bytes = shared_functions.read_file_bytes(filepath)
            file_content_str = file_content_bytes.decode('utf-8', errors='surrogateescape')
            file_content_str = file_content_str.replace('\r\n', '\n') # Normalize line endings.
        except OSError as e:
            raise OSError(f"Error reading file: {e}")
        if itemtype.lower() in ('container', 'table', 'sector', 'cell'):
            if address is None:
                raise ValueError(f"address is required for itemtype '{itemtype}'.")
            if not isinstance(address, list):
                raise TypeError("address must be a list.")
        if itemtype.lower() == 'container':
            if len(address) != 2:
                raise ValueError("address for container must contain containerid and name.")
            containerid = address[0]
            name = address[1]
            if not isinstance(containerid, int):
                raise TypeError("containerid must be an int")
            if not isinstance(name, str):
                raise TypeError("name must be a str")
            search_string = f'$<{containerid},{name}>'
            return search_string in file_content_str
        elif itemtype.lower() == 'table':
            if len(address) != 2:
                raise ValueError("address for table must contain tableid and name.")
            tableid = address[0]
            name = address[1]
            if not isinstance(tableid, int):
                raise TypeError("tableid must be an int")
            if not isinstance(name, str):
                raise TypeError("name must be a str")
            search_string = f'&<{tableid}^{name}>'
            return search_string in file_content_str
        elif itemtype.lower() == 'sector':
            if len(address) != 2:
                raise ValueError("address for sector must contain containerid and sectorid.")
            containerid = address[0]
            sectorid = address[1]
            if not isinstance(containerid, int):
                raise TypeError("containerid must be an int")
            if not isinstance(sectorid, int):
                raise TypeError("sectorid must be an int")
            search_string = f'!<[{containerid},{sectorid}],'
            return search_string in file_content_str
        elif itemtype.lower() == 'cell':
            if len(address) != 3:
                raise ValueError("address for cell must contain tableid, columnid, and rowid.")
            tableid = address[0]
            columnid = address[1]
            rowid = address[2]
            if not isinstance(tableid, int):
                raise TypeError("tableid must be an int")
            if not isinstance(columnid, int):
                raise TypeError("columnid must be an int")
            if not isinstance(rowid, int):
                raise TypeError("rowid must be an int")
            search_string = f'~<[{tableid};{columnid}?{rowid}],'
            return search_string in file_content_str
        else:
            return False
other = OtherClass()