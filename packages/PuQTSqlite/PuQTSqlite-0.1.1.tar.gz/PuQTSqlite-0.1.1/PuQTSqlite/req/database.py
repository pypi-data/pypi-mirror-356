import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime
from material_calc import calc_material

class Database:
    def __init__(self):
        self.db_file = "partners.db"
        # Создаем базу данных и заполняем данными только если она не существует
        if not os.path.exists(self.db_file):
            self._create_tables()
            self._insert_data()
    
    def _get_connection(self):
        return sqlite3.connect(self.db_file)
    
    def _create_tables(self):
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Создание таблицы типов материалов
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS типы_материалов (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                наименование TEXT NOT NULL,
                процент_брака REAL NOT NULL
            )
        ''')
        
        # Создание таблицы типов продукции
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS типы_продукции (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                наименование TEXT NOT NULL,
                коэффициент REAL NOT NULL
            )
        ''')
        
        # Создание таблицы продукции
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS продукция (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                тип_продукции TEXT NOT NULL,
                наименование TEXT NOT NULL,
                артикул TEXT NOT NULL,
                минимальная_стоимость REAL NOT NULL,
                FOREIGN KEY (тип_продукции) REFERENCES типы_продукции(наименование)
            )
        ''')
        
        # Создание таблицы партнеров
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS партнеры (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                тип_партнера TEXT NOT NULL,
                наименование_компании TEXT NOT NULL,
                фио_директора TEXT NOT NULL,
                email TEXT NOT NULL,
                телефон TEXT NOT NULL,
                адрес TEXT NOT NULL,
                инн TEXT NOT NULL,
                рейтинг INTEGER NOT NULL
            )
        ''')
        
        # Создание таблицы истории продаж
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS история_продаж (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                наименование_продукции TEXT NOT NULL,
                наименование_компании TEXT NOT NULL,
                количество INTEGER NOT NULL,
                дата_продажи DATE NOT NULL,
                FOREIGN KEY (наименование_продукции) REFERENCES продукция(наименование),
                FOREIGN KEY (наименование_компании) REFERENCES партнеры(наименование_компании)
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def _insert_data(self):
        conn = self._get_connection()
        cursor = conn.cursor()
        base_dir = os.path.dirname(os.path.abspath(__file__))

        # Helper function to read XLSX data
        def read_xlsx(file_name):
            file_path = os.path.join(base_dir, '..', file_name)
            workbook = load_workbook(file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            print(f"Заголовки в файле {file_name}:")
            print(headers)
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row): # Skip empty rows
                    row_dict = {}
                    for header, value in zip(headers, row):
                        row_dict[header] = value
                    data.append(row_dict)
            return data

        # Import material types
        material_types_data = read_xlsx('Material_type_import.xlsx')
        material_types_to_insert = []
        for row in material_types_data:
            percent_str = str(row['Процент брака материала']).replace(',', '.').strip('%')
            percent_value = float(percent_str) / 100
            material_types_to_insert.append((row['Тип материала'], percent_value))
        cursor.executemany('''
            INSERT INTO типы_материалов (наименование, процент_брака)
            VALUES (?, ?)
        ''', material_types_to_insert)

        # Import product types
        product_types_data = read_xlsx('Product_type_import.xlsx')
        product_types_to_insert = []
        for row in product_types_data:
            if row['Тип продукции']:
                coefficient = str(row['Коэффициент типа продукции']).replace(',', '.')
                product_types_to_insert.append((row['Тип продукции'], float(coefficient)))
        cursor.executemany('''
            INSERT INTO типы_продукции (наименование, коэффициент)
            VALUES (?, ?)
        ''', product_types_to_insert)

        # Import products
        products_data = read_xlsx('Products_import.xlsx')
        products_to_insert = []
        for row in products_data:
            min_cost = str(row['Минимальная стоимость для партнера']).replace(',', '.')
            products_to_insert.append((
                row['Тип продукции'],
                row['Наименование продукции'],
                row['Артикул'],
                float(min_cost)
            ))
        cursor.executemany('''
            INSERT INTO продукция (тип_продукции, наименование, артикул, минимальная_стоимость)
            VALUES (?, ?, ?, ?)
        ''', products_to_insert)

        # Import partners
        partners_data = read_xlsx('Partners_import.xlsx')
        partners_to_insert = []
        for row in partners_data:
            partners_to_insert.append((
                row['Тип партнера'],
                row['Наименование партнера'],
                row['Директор'],
                row['Электронная почта партнера'],
                row['Телефон партнера'],
                row['Юридический адрес партнера'],
                row['ИНН'],
                int(row['Рейтинг'])
            ))
        cursor.executemany('''
            INSERT INTO партнеры (тип_партнера, наименование_компании, фио_директора, email, телефон, адрес, инн, рейтинг)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', partners_to_insert)

        # Import sales history
        sales_history_data = read_xlsx('Partner_products_import.xlsx')
        sales_history_to_insert = []
        for row in sales_history_data:
            date_obj = row['Дата продажи']
            if isinstance(date_obj, datetime):
                formatted_date = date_obj.strftime('%Y-%m-%d')
            else:
                try:
                    if '.' in str(date_obj):
                        formatted_date = datetime.strptime(str(date_obj), '%d.%m.%Y').strftime('%Y-%m-%d')
                    else:
                        formatted_date = str(date_obj)
                except ValueError:
                    formatted_date = str(date_obj)

            sales_history_to_insert.append((
                row['Продукция'],
                row['Наименование партнера'],
                int(row['Количество продукции']),
                formatted_date
            ))
        cursor.executemany('''
            INSERT INTO история_продаж (наименование_продукции, наименование_компании, количество, дата_продажи)
            VALUES (?, ?, ?, ?)
        ''', sales_history_to_insert)

        conn.commit()
        conn.close()
    
    def get_partners(self):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM партнеры')
        partners = cursor.fetchall()
        conn.close()
        return partners
    
    def get_partner_history(self, partner_id):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT DISTINCT
                и.наименование_продукции,
                п.артикул,
                и.количество,
                и.дата_продажи,
                ROUND(п.минимальная_стоимость * и.количество, 2) as общая_сумма,
                CASE 
                    WHEN п.минимальная_стоимость * и.количество > 300000 THEN 15
                    WHEN п.минимальная_стоимость * и.количество > 50000 THEN 10
                    WHEN п.минимальная_стоимость * и.количество > 10000 THEN 5
                    ELSE 0
                END as процент_скидки
            FROM история_продаж и
            JOIN продукция п ON и.наименование_продукции = п.наименование
            JOIN партнеры пар ON и.наименование_компании = пар.наименование_компании
            WHERE пар.id = ?
            ORDER BY и.дата_продажи DESC
        ''', (partner_id,))
        history = cursor.fetchall()
        conn.close()
        return history
    
    def add_partner(self, partner_data):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO партнеры (
                тип_партнера,
                наименование_компании,
                фио_директора,
                email,
                телефон,
                адрес,
                инн,
                рейтинг
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', partner_data)
        conn.commit()
        conn.close()
        return cursor.lastrowid
    
    def update_partner(self, partner_id, partner_data):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE партнеры SET
                тип_партнера = ?,
                наименование_компании = ?,
                фио_директора = ?,
                email = ?,
                телефон = ?,
                адрес = ?,
                инн = ?,
                рейтинг = ?
            WHERE id = ?
        ''', partner_data + (partner_id,))
        conn.commit()
        conn.close()
    
    def delete_partner(self, partner_id):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM партнеры WHERE id = ?', (partner_id,))
        conn.commit()
        conn.close()
    
    def add_sale(self, partner_id, product_name, article, quantity, amount, discount=0):
        conn = self._get_connection()
        cursor = conn.cursor()
        date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
        INSERT INTO история_продаж (id_партнера, id_продукции, количество, 
                                  дата_продажи, общая_сумма, процент_скидки)
        VALUES (?, ?, ?, ?, ?, ?)
        ''', (partner_id, product_name, article, quantity, date, amount, discount))
        conn.commit()
        conn.close()
    
    def calculate_discount(self, total_amount):
        if total_amount >= 300000:
            return 15.0
        elif total_amount >= 50000:
            return 10.0
        elif total_amount >= 10000:
            return 5.0
        return 0.0

    def get_product_types(self):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id, наименование FROM типы_продукции')
        res = cursor.fetchall()
        conn.close()
        return res

    def get_material_types(self):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id, наименование FROM типы_материалов')
        res = cursor.fetchall()
        conn.close()
        return res

    def calc_material(self, prod_type_id, mat_type_id, qty, param1, param2):
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT коэффициент FROM типы_продукции WHERE id=?', (prod_type_id,))
        prod_row = cursor.fetchone()
        cursor.execute('SELECT процент_брака FROM типы_материалов WHERE id=?', (mat_type_id,))
        mat_row = cursor.fetchone()
        conn.close()
        if not prod_row or not mat_row:
            return -1
        return calc_material(prod_row[0], mat_row[0], qty, param1, param2)

    def get_partner_discount(self, partner_id):
        history = self.get_partner_history(partner_id)
        if not history:
            return 0
        total_qty = sum([h[2] for h in history])
        if total_qty > 300000:
            return 15
        elif total_qty > 50000:
            return 10
        elif total_qty > 10000:
            return 5
        else:
            return 0 