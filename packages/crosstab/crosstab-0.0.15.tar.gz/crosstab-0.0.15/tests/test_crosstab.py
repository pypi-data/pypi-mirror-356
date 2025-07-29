#!/usr/bin/env python

import logging
import tempfile
from pathlib import Path

import openpyxl
import pytest

from crosstab.crosstab import Crosstab

logger = logging.getLogger(__name__)

# Sample data for tests
CSV_CONTENT = """header1,header2,header3,value,unit
A,1,2018,10,%
A,1,2019,20,%
B,2,2018,30,%
B,2,2019,40,%
"""


@pytest.fixture(scope="session")
def global_variables():
    """Set global variables for the test session."""
    try:
        return {
            "SAMPLE_DATA_1": Path(__file__).parent / "data/sample1.csv",
            "SAMPLE_DATA_2": Path(__file__).parent / "data/sample2.csv",
        }
    except Exception:
        return None


def test_crosstab_init(global_variables):
    assert 1 == 1


@pytest.fixture
def temp_csv_file():
    """Fixture to create a temporary CSV file"""
    with tempfile.NamedTemporaryFile(mode="w+", suffix=".csv", delete=False) as f:
        f.write(CSV_CONTENT)
        f.seek(0)
        yield Path(f.name)
    Path(f.name).unlink()  # Clean up


@pytest.fixture
def temp_xlsx_file():
    """Fixture to create a temporary XLSX file"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        yield Path(f.name)
    Path(f.name).unlink()  # Clean up


def test_validate_args(temp_csv_file, temp_xlsx_file):
    """Test the validation of arguments."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2",),
        value_cols=("value", "unit"),
    )
    assert crosstab.incsv == temp_csv_file
    assert crosstab.outxlsx == temp_xlsx_file


def test_invalid_args_missing_file():
    """Test with missing CSV file"""
    with pytest.raises(ValueError, match="Input file .* does not exist."):
        Crosstab(
            incsv=Path("missing.csv"),
            outxlsx=Path("output.xlsx"),
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=("value", "unit"),
        )


def test_invalid_args_empty_file():
    """Test with an empty CSV file."""
    # Create an empty temporary CSV file
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
        temp_csv = Path(f.name)
    with pytest.raises(ValueError, match="Input file .* is empty."):
        Crosstab(
            incsv=temp_csv,
            outxlsx=Path("output.xlsx"),
            row_headers=("header1",),
            col_headers=("header2",),
            value_cols=("value", "unit"),
        )
    temp_csv.unlink()


def test_csv_to_sqlite(temp_csv_file):
    """Test the conversion of a CSV file to SQLite."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=Path("output.xlsx"),
        row_headers=("header1",),
        col_headers=("header2",),
        value_cols=("value", "unit"),
    )
    conn = crosstab._csv_to_sqlite()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='data';")
    row = cursor.fetchone()
    assert row is not None
    assert row[0] == "data"
    conn.close()


def test_crosstab_creation(temp_csv_file, temp_xlsx_file):
    """Test the creation of a crosstab file."""
    crosstab = Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value", "unit"),
        keep_sqlite=True,
        keep_src=True,
    )
    crosstab.crosstab()
    assert temp_xlsx_file.exists()
    # Test that the xlsx file has 3 sheets
    wb = openpyxl.load_workbook(temp_xlsx_file)
    assert len(wb.sheetnames) == 3


def test_crosstab_rows_single_value_column(temp_csv_file, temp_xlsx_file):
    """Test that values are correctly placed in the crosstab with one value column."""
    Crosstab(
        incsv=temp_csv_file,
        outxlsx=temp_xlsx_file,
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value",),
        keep_sqlite=False,
        keep_src=True,
    ).crosstab()
    wb = openpyxl.load_workbook(temp_xlsx_file)
    ws = wb["Crosstab"]
    # Check the row headers
    assert ws["A1"].value == "header2"
    assert ws["A2"].value == "header3"
    assert ws["A3"].value == "header1"
    assert ws["A4"].value == "A"
    assert ws["A5"].value == "B"
    # Check the column headers
    assert ws["B1"].value == "1"
    assert ws["C1"].value == "1"
    assert ws["D1"].value == "2"
    assert ws["E1"].value == "2"
    assert ws["B2"].value == "2018"
    assert ws["C2"].value == "2019"
    assert ws["D2"].value == "2018"
    assert ws["E2"].value == "2019"
    assert ws["B3"].value == "value"
    assert ws["C3"].value == "value"
    assert ws["D3"].value == "value"
    assert ws["E3"].value == "value"
    # Check the values
    assert ws["B4"].value == "10"
    assert ws["C4"].value == "20"
    assert ws["D4"].value is None
    assert ws["E4"].value is None
    assert ws["B5"].value is None
    assert ws["C5"].value is None
    assert ws["D5"].value == "30"
    assert ws["E5"].value == "40"
    wb.close()


def test_crosstab_rows_multi_value_column(temp_csv_file, temp_xlsx_file):
    """Test that values are correctly placed in the crosstab with multiple value columns."""
    Crosstab(
        incsv=temp_csv_file,
        outxlsx=Path(temp_xlsx_file),
        row_headers=("header1",),
        col_headers=("header2", "header3"),
        value_cols=("value", "unit"),
        keep_sqlite=False,
        keep_src=True,
    ).crosstab()
    wb = openpyxl.load_workbook(temp_xlsx_file)
    ws = wb["Crosstab"]
    # Check the row headers
    assert ws["A1"].value == "header2"
    assert ws["A2"].value == "header3"
    assert ws["A3"].value == "header1"
    assert ws["A4"].value == "A"
    assert ws["A5"].value == "B"
    # Check the column headers
    assert ws["B1"].value == "1"
    assert ws["C1"].value is None
    assert ws["D1"].value == "1"
    assert ws["E1"].value is None
    assert ws["F1"].value == "2"
    assert ws["G1"].value is None
    assert ws["H1"].value == "2"
    assert ws["I1"].value is None
    assert ws["B2"].value == "2018"
    assert ws["C2"].value is None
    assert ws["D2"].value == "2019"
    assert ws["E2"].value is None
    assert ws["F2"].value == "2018"
    assert ws["G2"].value is None
    assert ws["H2"].value == "2019"
    assert ws["I2"].value is None
    assert ws["B3"].value == "value"
    assert ws["C3"].value == "unit"
    assert ws["D3"].value == "value"
    assert ws["E3"].value == "unit"
    assert ws["F3"].value == "value"
    assert ws["G3"].value == "unit"
    assert ws["H3"].value == "value"
    assert ws["I3"].value == "unit"
    # Check the values
    assert ws["B4"].value == "10"
    assert ws["C4"].value == "%"
    assert ws["D4"].value == "20"
    assert ws["E4"].value == "%"
    assert ws["F4"].value is None
    assert ws["G4"].value is None
    assert ws["H4"].value is None
    assert ws["I4"].value is None
    assert ws["B5"].value is None
    assert ws["C5"].value is None
    assert ws["D5"].value is None
    assert ws["E5"].value is None
    assert ws["F5"].value == "30"
    assert ws["G5"].value == "%"
    assert ws["H5"].value == "40"
    assert ws["I5"].value == "%"
    wb.close()
