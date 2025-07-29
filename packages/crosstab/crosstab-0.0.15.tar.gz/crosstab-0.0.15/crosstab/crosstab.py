#!/usr/bin/env python

from __future__ import annotations

import argparse
import csv
import datetime
import getpass
import logging
import sqlite3
import sys
import traceback
from collections.abc import Generator
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Color, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.utils.cell import get_column_letter

__title__ = "crosstab"
__author__ = "Caleb Grant"
__url__ = "https://github.com/geocoug/crosstab"
__author_email__ = "grantcaleb22@gmail.com"
__license__ = "GNU GPLv3"
__version__ = "0.0.15"
__description__ = "Rearrange data from a normalized CSV format to a crosstabulated format, with styling."

logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[logging.NullHandler()],
)
logger = logging.getLogger(__name__)

# Row number to start writing the crosstab table. Must be > 0.
XTAB_START_ROW = 1
# Column number to start writing the crosstab table. Must be > 0.
XTAB_START_COL = 1

"""
Excel styling using named styles.

See: https://openpyxl.readthedocs.io/en/stable/styles.html
"""
# Styling for the primary header rows (row headers and column headers)
PRIMARY_HEADER_STYLE = NamedStyle(name="PRIMARY_HEADER_STYLE")
PRIMARY_HEADER_STYLE.font = Font(bold=True, size=12)
PRIMARY_HEADER_STYLE.fill = PatternFill(patternType="solid", fgColor=Color("D9D9D9"))
PRIMARY_HEADER_STYLE.alignment = Alignment(horizontal="center", vertical="center")
PRIMARY_HEADER_STYLE.border = Border(
    left=Side(border_style="thin", color=Color("000000")),
    right=Side(border_style="thin", color=Color("000000")),
    top=Side(border_style="thin", color=Color("000000")),
    bottom=Side(border_style="thin", color=Color("000000")),
)

# Styling for the secondary header rows (value columns)
SECONDARY_HEADER_STYLE = NamedStyle(name="SECONDARY_HEADER_STYLE")
SECONDARY_HEADER_STYLE.font = Font(bold=True, size=12)
SECONDARY_HEADER_STYLE.fill = PatternFill(patternType="solid", fgColor=Color("F2F2F2"))
SECONDARY_HEADER_STYLE.alignment = Alignment(horizontal="center", vertical="center")
SECONDARY_HEADER_STYLE.border = Border(
    left=Side(border_style="thin", color=Color("000000")),
    right=Side(border_style="thin", color=Color("000000")),
    top=Side(border_style="thin", color=Color("000000")),
    bottom=Side(border_style="thin", color=Color("000000")),
)

# Styling for the data cells
DATA_STYLE = NamedStyle(name="DATA_STYLE")
DATA_STYLE.font = Font(bold=False, size=12)
DATA_STYLE.fill = PatternFill(patternType="solid", fgColor=Color("ffffff"))
DATA_STYLE.alignment = Alignment(horizontal="left", vertical="center")
DATA_STYLE.border = Border(
    left=Side(border_style="thin", color=Color("000000")),
    right=Side(border_style="thin", color=Color("000000")),
    top=Side(border_style="thin", color=Color("000000")),
    bottom=Side(border_style="thin", color=Color("000000")),
)

# Styling for the title text
TITLE_TEXT_STYLE = NamedStyle(name="TITLE_TEXT_STYLE")
TITLE_TEXT_STYLE.font = Font(name="Arial", bold=True, size=11, color=Color("005782"))
TITLE_TEXT_STYLE.alignment = Alignment(horizontal="center", vertical="center")

# Styling for the README sheet - items
METADATA_ITEM_STYLE = NamedStyle(name="METADATA_ITEM_STYLE")
METADATA_ITEM_STYLE.font = Font(bold=True, size=12)
METADATA_ITEM_STYLE.alignment = Alignment(horizontal="right", vertical="center")
METADATA_ITEM_STYLE.border = Border(
    left=Side(border_style="thin", color=Color("000000")),
    right=Side(border_style="thin", color=Color("000000")),
    top=Side(border_style="thin", color=Color("000000")),
    bottom=Side(border_style="thin", color=Color("000000")),
)

# Styling for the README sheet - values
METADATA_VALUE_STYLE = NamedStyle(name="METADATA_VALUE_STYLE")
METADATA_VALUE_STYLE.font = Font(bold=False, size=12)
METADATA_VALUE_STYLE.alignment = Alignment(horizontal="left", vertical="center")
METADATA_VALUE_STYLE.border = Border(
    left=Side(border_style="thin", color=Color("000000")),
    right=Side(border_style="thin", color=Color("000000")),
    top=Side(border_style="thin", color=Color("000000")),
    bottom=Side(border_style="thin", color=Color("000000")),
)


class Crosstab:
    def __init__(
        self: Crosstab,
        incsv: Path,
        row_headers: tuple,
        col_headers: tuple,
        value_cols: tuple,
        outxlsx: Path | None = None,
        keep_sqlite: bool = False,
        keep_src: bool = False,
    ) -> None:
        """Create a crosstab table from a normalized CSV file.

        Args:
            incsv (Path): Path to the input CSV file.
            outxlsx (Path, optional): Path to the output XLSX file. The output file will contain at a minimum two sheets: one containing metadata about the crosstab and one containing the crosstab table. If the keep_src argument is True, the output file will contain a third sheet with the source data. If no output file is specified, the output will be written in the same directory as the input file, with the same name as the input file, appended with "_crosstab" and an XLSX extension. Defaults to None.
            row_headers (tuple): Tuple of one or more column names to use as row headers. Unique values of these columns will appear at the beginning of every output line.
            col_headers (tuple): Tuple of one or more column names to use as column headers in the output. A crosstab column (or columns) will be created for every unique combination of values of these fields in the input.
            value_cols (tuple): Tuple of one or more column names with values to be used to fill the cells of the cross-table. If n columns names are specified, then there will be n columns in the output table for each of the column headers corresponding to values of the -c argument. The column names specified with the -v argument will be appended to the output column headers created from values of the -c argument. There should be only one value of the -v column(s) for each combination of the -r and -c columns; if there is more than one, a warning will be printed and only the first value will appear in the output. (That is, values are not combined in any way when there are multiple values for each output cell.)
            keep_sqlite (bool, optional): Keep the temporary SQLite database file. The default is to delete it after the output file is created. The SQLite file is created in the same directory as the input file with the name of the input file (but with a .sqlite extension) and a single table named 'data'. Defaults to False.
            keep_src (bool, optional): Keep a sheet with the source data in the output file. The sheet will be named 'Source Data'. Defaults to False.

        Raises:
            ValueError: Raised if the input file does not exist, is not a file, is empty, is not a CSV file, or if the row_headers, col_headers, or value_cols are not specified. Also raised if the output file does not have an XLSX extension.

        Example:

        If you have a CSV file with the following data:

        ```csv
        location,sample,cas_rn,parameter,concentration,units
        Loc1,Samp1,7440-66-6,Zinc,1.0,mg/L
        Loc1,Samp1,7439-89-6,Iron,2.7,mg/L
        Loc2,Samp2,7440-66-6,Zinc,8.0,mg/L
        Loc2,Samp2,7439-89-6,Iron,3.23,mg/L
        ```

        The code...

        ```python
        from pathlib import Path
        from crosstab import Crosstab

        Crosstab(
            incsv=Path("input.csv"),
            outxlsx=Path("output.xlsx"),
            row_headers=("location", "sample"),
            col_headers=("cas_rn", "parameter"),
            value_cols=("concentration", "units"),
            keep_sqlite=True,
            keep_src=True,
        ).crosstab()
        ```

        ...will produce a crosstab table with the following structure:

        ```txt
        +----------------------------------------------------------------------+
        │          │   cas_rn  │       7440-66-6       │       7439-89-6       │
        │          │ --------- │ --------------------- │ --------------------- │
        │          │ parameter │          Zinc         │          Iron         │
        │ -------- │ --------- │ --------------------- │ --------------------- │
        │ location │  sample   │ concentration │ units │ concentration │ units │
        │==========│===========│===============│=======│===============│=======│
        │ Loc1     │ Samp1     │ 1.0           │ mg/L  │ 2.7           │ mg/L  │
        │ Loc2     │ Samp2     │ 8.0           │ mg/L  │ 3.23          │ mg/L  │
        +----------------------------------------------------------------------+
        ```
        """  # noqa: E501
        self.incsv = incsv
        if not outxlsx:
            outxlsx = incsv.with_name(incsv.stem + "_crosstab.xlsx")
        self.outxlsx = outxlsx
        self.row_headers = row_headers
        self.col_headers = col_headers
        self.value_cols = value_cols
        self.keep_sqlite = keep_sqlite
        self.keep_src = keep_src
        logger.debug(self)
        self._validate_args()
        with open(self.incsv) as f:
            self.dialect = csv.Sniffer().sniff(f.readline())
        self.csv_columns = next(self._csv_reader())
        self.csv_reader = self._csv_reader()
        self._validate_csv_headers()
        self.conn = self._csv_to_sqlite()

    def __repr__(self: Crosstab) -> str:
        return f"Crosstab(incsv={self.incsv!r}, outxlsx={self.outxlsx!r}, row_headers={self.row_headers!r}, col_headers={self.col_headers!r}, value_cols={self.value_cols!r}, keep_sqlite={self.keep_sqlite!r})"  # noqa: E501

    def _validate_args(self: Crosstab) -> None:
        """Validate arguments passed to the Crosstab class."""
        if not self.incsv.exists():
            raise ValueError(f"Input file {self.incsv} does not exist.")
        if not self.incsv.is_file():
            raise ValueError(f"Input file {self.incsv} is not a file.")
        if not self.incsv.stat().st_size:
            raise ValueError(f"Input file {self.incsv} is empty.")
        if not self.incsv.suffix == ".csv":
            raise ValueError(f"Input file {self.incsv} is not a CSV file.")
        if not self.outxlsx.suffix == ".xlsx":
            raise ValueError("Output file must have an XLSX extension.")
        if not self.row_headers:
            raise ValueError("No row headers specified.")
        if not self.col_headers:
            raise ValueError("No column headers specified.")
        if not self.value_cols:
            raise ValueError("No value columns specified.")

    def _csv_reader(self: Crosstab) -> Generator:
        """Read the CSV file and yield each row as a dictionary."""
        with open(self.incsv, newline="") as f:
            reader = csv.DictReader(f, dialect=self.dialect)
            yield from reader

    def _validate_csv_headers(self: Crosstab) -> None:
        """Validate all row_headers, col_headers, and value_cols exist in the CSV file."""
        bad_headers = []
        for header in self.row_headers + self.col_headers + self.value_cols:
            if header not in self.csv_columns:
                bad_headers.append(header)
        if bad_headers:
            raise ValueError(
                f"Headers not found in CSV file: {', '.join(bad_headers)}.",
            )

    def _csv_to_sqlite(self: Crosstab) -> sqlite3.Connection:
        """Convert the CSV file to a SQLite database.

        If the keep_sqlite attribute is True, the SQLite database file will be saved to disk. Otherwise, the database will be created in memory. The database will have a single table named 'data' with columns corresponding to the CSV file headers.
        """  # noqa: E501
        if self.keep_sqlite:
            sqlite_file = self.incsv.with_suffix(".sqlite")
            logger.info(f"Creating SQLite database file: {sqlite_file}.")
            if sqlite_file.exists():
                sqlite_file.unlink()
            conn = sqlite3.connect(sqlite_file)
        else:
            logger.debug("Creating in-memory SQLite database.")
            conn = sqlite3.connect(":memory:")
        logger.debug("Creating 'data' table in SQLite database.")
        with conn:
            cursor = conn.cursor()
            coldef = ", ".join([f'"{col}"' for col in self.csv_columns])
            cursor.execute(
                f"CREATE TABLE data ({coldef});",
            )
            cursor.executemany(
                f"INSERT INTO data VALUES ({', '.join(['?' for _ in self.csv_columns])});",
                (tuple(row.values()) for row in self.csv_reader),
            )
        return conn

    def crosstab(self: Crosstab) -> None:
        """Create a crosstab table from the input CSV file.

        The crosstab table will be written to the output XLSX file. The table will have row headers, column headers, and value columns as specified in the `row_headers`, `col_headers`, and `value_cols` arguments. The table will be written to a sheet named *Crosstab*. If the `keep_src` argument is `True`, a sheet named *Source Data* will be created with the source data from the input CSV file. A sheet named *README* will be created with metadata about the crosstab process. The metadata will include the creation time, user, script version, input file, output file, and SQLite file (if the `keep_sqlite` argument is `True`). Both the *README* and *Crosstab* sheets will be styled to make the table easier to read.
        """  # noqa: E501
        logger.info(f"Creating crosstab table from {self.incsv}.")
        # Get list of unique values for each row header
        logger.debug("Getting list of unique values for each row header.")
        with self.conn:
            cursor = self.conn.cursor()
            cursor.execute(f"SELECT DISTINCT {', '.join(self.row_headers)} FROM data;")
            row_header_vals = cursor.fetchall()

        # Check if there are multiple value columns for each row header key/column header key combination
        logger.debug("Checking for multiple value columns for each row header key.")
        with self.conn:
            sql = f"SELECT {', '.join(self.row_headers + self.col_headers)}, COUNT(*) FROM data GROUP BY {', '.join(self.row_headers + self.col_headers)} HAVING COUNT(*) > 1;"  # noqa: E501
            logger.debug(sql)
            cursor = self.conn.cursor()
            cursor.execute(sql)
            multiple_vals = cursor.fetchall()
            if multiple_vals:
                raise ValueError(
                    "Multiple values found for the row/column combination(s).",
                )

        # Create the workbook
        logger.debug("Initializing the workbook.")
        wb = openpyxl.Workbook()

        # Add a README sheet with metadata about the crosstab
        logger.debug("Creating README sheet.")
        readme = wb.active
        if not readme:
            readme = wb.create_sheet()
        readme.title = "README"
        metadata = {
            "Creation Time": datetime.datetime.now().isoformat(
                sep=" ",
                timespec="seconds",
            ),
            "User": getpass.getuser(),
            "Script Version": __version__,
            "Input File": self.incsv.resolve().as_posix(),
            "Output File": self.outxlsx.resolve().as_posix(),
            "SQLite File": (self.outxlsx.with_suffix(".sqlite").resolve().as_posix() if self.keep_sqlite else None),
        }
        readme.cell(row=1, column=1, value="Crosstab Metadata").style = TITLE_TEXT_STYLE
        readme.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        readme.cell(row=2, column=1, value="Item").style = PRIMARY_HEADER_STYLE
        readme.cell(row=2, column=2, value="Value").style = PRIMARY_HEADER_STYLE
        for i, (item, value) in enumerate(metadata.items(), start=1):
            readme.cell(row=i + 2, column=1, value=item).style = METADATA_ITEM_STYLE
            readme.cell(row=i + 2, column=2, value=value).style = METADATA_VALUE_STYLE
        # Auto size all columns
        for col in readme.columns:
            length = max(len(str(cell.value)) for cell in col)
            readme.column_dimensions[get_column_letter(col[0].column)].width = length

        # Create the crosstab sheet
        logger.debug("Creating crosstab sheet.")
        sheet = wb.create_sheet("Crosstab")

        # Get all of the unique column header values
        logger.debug("Getting list of unique values for each column header.")
        with self.conn:
            cursor = self.conn.cursor()
            cursor.execute(f"SELECT DISTINCT {', '.join(self.col_headers)} FROM data;")
            col_header_vals = cursor.fetchall()
        logger.debug("Writing crosstab table.")
        # Write the column headers. If multiple column headers are specified, write them in separate rows.
        # Space out each column by the number of value_cols specified.
        start_col = len(self.row_headers) + XTAB_START_COL  # Column index to start writing the column headers
        for i, col_header in enumerate(col_header_vals):
            for c, hdr in enumerate(col_header):
                if i == 0:
                    logger.debug(f"Writing column header row: {self.col_headers[c]}.")
                    # Label the header row
                    sheet.cell(
                        row=c + XTAB_START_ROW,
                        column=start_col - 1,
                        value=self.col_headers[c],
                    )
                    # Style the labels
                    sheet.cell(row=c + XTAB_START_ROW, column=start_col - 1).style = SECONDARY_HEADER_STYLE
                # Add the column header
                sheet.cell(row=c + XTAB_START_ROW, column=start_col + i, value=hdr)
                # Merge the column header cells
                sheet.merge_cells(
                    start_row=c + XTAB_START_ROW,
                    start_column=start_col + i,
                    end_row=c + XTAB_START_ROW,
                    end_column=start_col + i + len(self.value_cols) - 1,
                )
                # Apply the header style to the merged cells
                for col in range(start_col + i, start_col + i + len(self.value_cols)):
                    sheet.cell(row=c + XTAB_START_ROW, column=col).style = PRIMARY_HEADER_STYLE
            # Add the value columns below the column headers
            logger.debug(f"Writing value columns for column header row: {col_header}.")
            for j, value_col in enumerate(self.value_cols):
                sheet.cell(
                    row=len(self.col_headers) + XTAB_START_ROW,
                    column=start_col + i + j,
                    value=value_col,
                ).style = SECONDARY_HEADER_STYLE
            # If this is the last column header, reset the start_col index, otherwise, add the number of value columns
            if i == len(col_header_vals) - 1:
                start_col += len(self.value_cols) + 1
            else:
                start_col += len(self.value_cols) - 1

        # Write the row headers and value columns
        logger.debug(f"Writing row headers: {self.row_headers}.")
        start_row = len(self.col_headers) + XTAB_START_ROW
        for i in range(len(self.row_headers)):
            sheet.cell(
                row=start_row,
                column=i + XTAB_START_COL,
                value=self.row_headers[i],
            ).style = PRIMARY_HEADER_STYLE

        # Write the data
        logger.debug("Writing crosstab data.")
        with self.conn:
            cursor = self.conn.cursor()
            for i, row_header in enumerate(row_header_vals):
                # Write the row keys
                for h, hdr in enumerate(row_header):
                    sheet.cell(
                        row=start_row + i + 1,
                        column=h + XTAB_START_COL,
                        value=hdr,
                    ).style = DATA_STYLE
                # Write the data
                for j, col_header in enumerate(col_header_vals):
                    cols = ", ".join(self.value_cols)
                    where1 = " AND ".join(
                        [
                            "\"{}\" = '{}'".format(
                                self.row_headers[k],
                                row_header[k].replace("'", "''"),
                            )
                            for k in range(len(self.row_headers))
                        ],
                    )
                    where2 = " AND ".join(
                        [
                            "\"{}\" = '{}'".format(
                                self.col_headers[k],
                                col_header[k].replace("'", "''"),
                            )
                            for k in range(len(self.col_headers))
                        ],
                    )
                    sql = f"SELECT {cols} FROM data WHERE {where1} AND {where2};"
                    logger.debug(sql)
                    cursor.execute(sql)
                    if cursor.rowcount > 1:
                        raise ValueError(
                            f"Multiple values found for row/column combination: {row_header}, {col_header}",
                        )
                    data = cursor.fetchone()
                    if data:
                        for k, value in enumerate(data):
                            sheet.cell(
                                row=start_row + i + 1,
                                column=len(self.row_headers) + XTAB_START_COL + j * len(self.value_cols) + k,
                                value=value,
                            ).style = DATA_STYLE
                    else:
                        for k in range(len(self.value_cols)):
                            sheet.cell(
                                row=start_row + i + 1,
                                column=len(self.row_headers) + XTAB_START_COL + j * len(self.value_cols) + k,
                                value="",
                            ).style = DATA_STYLE

        # Turn on filter for the row headers
        filter_range = (
            get_column_letter(XTAB_START_COL)
            + str(start_row)
            + ":"
            + get_column_letter(XTAB_START_COL + len(self.row_headers) - 1)
            + str(start_row)
        )
        logger.debug(f"Applying filters to the row headers: {filter_range}.")
        sheet.auto_filter.ref = filter_range

        # Freeze the row headers
        logger.debug(f"Freezing the row headers: {filter_range}.")
        sheet.freeze_panes = sheet.cell(row=start_row + 1, column=XTAB_START_COL)

        # Auto size all columns
        logger.debug("Auto-sizing all columns.")
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].auto_size = True

        # Add a sheet with the source data
        if self.keep_src:
            logger.debug("Creating source data sheet.")
            src_data = wb.create_sheet(title="Source Data")
            with self.conn:
                cursor = self.conn.cursor()
                cursor.execute("SELECT * FROM data;")
                raw_data_headers = [description[0] for description in cursor.description]
                src_data.append(raw_data_headers)
                for row in cursor.fetchall():
                    src_data.append(row)

        logger.info(f"Saving output to {self.outxlsx}.")
        wb.save(self.outxlsx)


def clparser() -> argparse.ArgumentParser:
    """Command line interface for the upsert function."""
    parser = argparse.ArgumentParser(
        description=__description__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""notes:\n1. Column names should be specified in the same case as they appear in the input file. If column names contain spaces, they should be wrapped in quotes.\n2. The -f option creates a temporary file in the same directory as the output file.  This file has the same name as the input file, but an extension of '.sqlite'.\n3. There are no inherent limits to the number of rows or columns in the input or output files.\n4. Missing required arguments will result in an exception rather than an error message, whatever the error logging option.  If no error logging option is specified, then if there are multiple values per cell (the most likely data error), a single message will be printed on the console.""",  # noqa: E501
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {__version__}",
    )
    parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="suppress all console output",
    )
    parser.add_argument(
        "-d",
        "--debug",
        action="store_true",
        help="enable debug logging",
    )
    parser.add_argument(
        "-l",
        "--log",
        metavar="LOGFILE",
        dest="log",
        type=Path,
        help="log all output to a file",
    )
    parser.add_argument(
        "-k",
        "--keep-sqlite",
        action="store_true",
        help="keep the temporary SQLite database file. The default is to delete it after the output file is created. The SQLite file is created in the same directory as the output file with the name of the output file (but with a .sqlite extension) and a single table named 'data'.",  # noqa: E501
    )
    parser.add_argument(
        "-s",
        "--keep-src",
        action="store_true",
        help="keep a sheet with the source data in the output file. The default is to not include the source data in the output file.",  # noqa: E501
    )
    parser.add_argument(
        "-f",
        metavar="INPUT_CSV",
        dest="incsv",
        required=True,
        type=Path,
        help="input CSV file",
    )
    parser.add_argument(
        "-o",
        metavar="OUTPUT_XLSX",
        dest="outxlsx",
        required=False,
        type=Path,
        help="output XLSX file",
    )
    parser.add_argument(
        "-r",
        metavar="ROW_HEADERS",
        dest="row_headers",
        required=True,
        nargs="+",
        help="one or more column names to use as row headers. Unique values of these columns will appear at the beginning of every output line.",  # noqa: E501
    )
    parser.add_argument(
        "-c",
        metavar="COL_HEADERS",
        dest="col_headers",
        required=True,
        nargs="+",
        help="one or more column names to use as column headers in the output. A crosstab column (or columns) will be created for every unique combination of values of these fields in the input.",  # noqa: E501
    )
    parser.add_argument(
        "-v",
        metavar="VALUE_COLS",
        dest="value_cols",
        required=True,
        nargs="+",
        help="one or more column names with values to be used to fill the cells of the cross-table.  If n columns names are specified, then there will be n columns in the output table for each of the column headers corresponding to values of the -c argument.  The column names specified with the -v argument will be appended to the output column headers created from values of the -c argument.  There should be only one value of the -v column(s) for each combination of the -r and -c columns; if there is more than one, a warning will be printed and only the first value will appear in the output.  (That is, values are not combined in any way when there are multiple values for each output cell.)",  # noqa: E501
    )
    return parser


def cli() -> None:
    """Main command line entrypoint for the xtab function."""
    args = clparser().parse_args()
    if not args.quiet:
        logger.addHandler(logging.StreamHandler())
    if args.log:
        logger.addHandler(logging.FileHandler(args.log))
    if args.debug:
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter(
            "%(asctime)s %(name)s (%(lineno)d) %(levelname)s: %(message)s",
            datefmt="[%Y-%m-%d %H:%M:%S]",
        )
        for handler in logger.handlers:
            handler.setFormatter(formatter)
    try:
        Crosstab(
            incsv=args.incsv,
            outxlsx=args.outxlsx,
            row_headers=args.row_headers,
            col_headers=args.col_headers,
            value_cols=args.value_cols,
            keep_sqlite=args.keep_sqlite,
            keep_src=args.keep_src,
        ).crosstab()
    except SystemExit as x:
        sys.exit(x.code)
    except ValueError:
        strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
        lno = strace[0][1]
        src = strace[0][3]
        logger.error(
            f"ValueError on line {lno}: {sys.exc_info()[1]}",
        )
        sys.exit(1)
    except Exception:
        strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
        lno = strace[0][1]
        src = strace[0][3]
        logger.error(
            f"Uncaught exception {sys.exc_info()[0]!s} ({sys.exc_info()[1]}) on line {lno} ({src}).",
        )
        sys.exit(1)


if __name__ == "__main__":
    cli()
